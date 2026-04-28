"""
Модул за обработка на изходни данни
Създава интерактивна карта, Excel файлове и чартове за анализ
"""

import folium
from branca.element import MacroElement, Template
import pandas as pd
import requests
import json
import html
import math
from typing import List, Dict, Tuple, Optional
import os
import logging
from datetime import datetime
from config import get_config, OutputConfig, RoutingEngine, is_location_in_center_zone
from cvrp_solver import CVRPSolution, Route
from warehouse_manager import WarehouseAllocation
from input_handler import Customer
from osrm_client import get_distance_matrix_from_central_cache

try:
    from folium.plugins import PolyLineTextPath
except ImportError:
    PolyLineTextPath = None

# OpenPyXL imports за Excel стилове
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import matplotlib
    matplotlib.use('Agg')  # Без GUI
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

logger = logging.getLogger(__name__)


def _normalize_output_file_path(path_value: str, default_filename: str) -> str:
    """Normalize output file paths.

    If the config contains a directory instead of a file path, append a default
    filename so the application does not fail with PermissionError.
    """
    normalized = os.path.normpath(path_value)
    _, extension = os.path.splitext(normalized)
    if extension:
        return normalized
    return os.path.join(normalized, default_filename)


def _append_run_date_to_filename(file_path: str, run_date: Optional[str] = None) -> str:
    """Append optimizer run date before the file extension."""
    date_stamp = run_date or datetime.now().strftime("%Y-%m-%d")
    directory, filename = os.path.split(file_path)
    stem, extension = os.path.splitext(filename)
    if not extension or stem.endswith(f"_{date_stamp}"):
        return file_path
    return os.path.join(directory, f"{stem}_{date_stamp}{extension}")

# Настройки за различните типове превозни средства
VEHICLE_SETTINGS = {
    'internal_bus': {
        'color': 'blue',
        'icon': 'bus',
        'prefix': 'fa',
        'name': 'Вътрешен автобус'
    },
    'center_bus': {
        'color': 'red', 
        'icon': 'building',
        'prefix': 'fa',
        'name': 'Централен автобус'
    },
    'external_bus': {
        'color': 'red',
        'icon': 'truck',
        'prefix': 'fa', 
        'name': 'Външен автобус'
    },
    'vratza_bus': {
        'color': 'green',
        'icon': 'car',
        'prefix': 'fa',
        'name': 'Враца автобус'
    }
}

# Цветове за всеки отделен автобус
BUS_COLORS = [
    '#FF0000',  # Червен
    '#00FF00',  # Зелен  
    '#0000FF',  # Син
    '#FFFF00',  # Жълт
    '#FF00FF',  # Магента
    '#00FFFF',  # Циан
    '#FFA500',  # Оранжев
    '#800080',  # Лилав
    '#008000',  # Тъмно зелен
    '#000080',  # Тъмно син
    '#800000',  # Бордо
    '#808000',  # Маслинен
    '#FF69B4',  # Розов
    '#32CD32',  # Лайм зелен
    '#8A2BE2',  # Синьо виолетов
    '#FF4500',  # Червено оранжев
    '#2E8B57',  # Морско зелен
    '#4682B4',  # Стоманено син
    '#D2691E',  # Шоколадов
    '#DC143C'   # Тъмно червен
]


class SingleRouteCustomerPanel(MacroElement):
    """Leaflet control with clickable customer list for single route maps."""

    _template = Template(
        """
        {% macro script(this, kwargs) %}
        (function() {
            const map = {{ this._parent.get_name() }};
            const panelHtml = {{ this.panel_html|safe }};
            const clients = {{ this.clients|safe }};
            const css = {{ this.css|safe }};

            if (!document.getElementById("route-client-panel-style")) {
                const style = document.createElement("style");
                style.id = "route-client-panel-style";
                style.textContent = css;
                document.head.appendChild(style);
            }

            const panel = L.control({ position: "topright" });
            panel.onAdd = function() {
                const container = L.DomUtil.create("div", "route-client-control");
                container.innerHTML = `
                    <button type="button" class="route-client-toggle">Клиенти</button>
                    <div class="route-client-panel route-client-panel-hidden">${panelHtml}</div>
                `;
                L.DomEvent.disableClickPropagation(container);
                L.DomEvent.disableScrollPropagation(container);
                return container;
            };
            panel.addTo(map);

            function setPanelOpen(isOpen) {
                const panelEl = document.querySelector(".route-client-panel");
                const buttonEl = document.querySelector(".route-client-toggle");
                if (!panelEl || !buttonEl) return;
                panelEl.classList.toggle("route-client-panel-hidden", !isOpen);
                buttonEl.classList.toggle("route-client-toggle-hidden", isOpen);
            }

            function openClient(index) {
                const item = clients[index];
                if (!item) return;
                const marker = window[item.markerName];
                if (!marker) return;
                map.setView([item.lat, item.lng], Math.max(map.getZoom(), 15), { animate: true });
                marker.openPopup();
            }

            setTimeout(function() {
                const toggle = document.querySelector(".route-client-toggle");
                const closeButton = document.querySelector(".route-client-close");
                if (toggle) {
                    toggle.addEventListener("click", function() {
                        setPanelOpen(true);
                    });
                }
                if (closeButton) {
                    closeButton.addEventListener("click", function(event) {
                        event.preventDefault();
                        setPanelOpen(false);
                    });
                }

                document.querySelectorAll(".route-client-card").forEach(function(row) {
                    row.addEventListener("click", function(event) {
                        if (event.target.closest("a")) return;
                        openClient(Number(row.dataset.clientIndex));
                    });
                    row.addEventListener("keydown", function(event) {
                        if (event.key === "Enter" || event.key === " ") {
                            event.preventDefault();
                            openClient(Number(row.dataset.clientIndex));
                        }
                    });
                });
            }, 0);
        })();
        {% endmacro %}
        """
    )

    def __init__(self, panel_html: str, clients: List[Dict[str, object]], css: str):
        super().__init__()
        self._name = "SingleRouteCustomerPanel"
        self.panel_html = json.dumps(panel_html, ensure_ascii=False)
        self.clients = json.dumps(clients, ensure_ascii=False)
        self.css = json.dumps(css, ensure_ascii=False)


class GoogleMapDocument:
    """Small save-compatible wrapper for generated Google Maps HTML."""

    def __init__(self, html_content: str):
        self.html_content = html_content

    def save(self, file_path: str) -> None:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(self.html_content)


class StreetViewPicker(MacroElement):
    """Leaflet control that opens Google Street View for a clicked map point."""

    _template = Template(
        """
        {% macro script(this, kwargs) %}
        (function() {
            const map = {{ this._parent.get_name() }};
            const css = {{ this.css|safe }};

            if (!document.getElementById("street-view-picker-style")) {
                const style = document.createElement("style");
                style.id = "street-view-picker-style";
                style.textContent = css;
                document.head.appendChild(style);
            }

            const picker = L.control({ position: "topleft" });
            picker.onAdd = function() {
                const container = L.DomUtil.create("div", "street-view-picker-control");
                container.innerHTML = `
                    <button type="button" class="street-view-picker-button" title="Отвори Google Street View от избрана точка">
                        <span class="street-view-picker-icon" aria-hidden="true"></span>
                        <span>Street View</span>
                    </button>
                `;
                L.DomEvent.disableClickPropagation(container);
                L.DomEvent.disableScrollPropagation(container);
                return container;
            };
            picker.addTo(map);

            function streetViewUrl(lat, lng) {
                return "https://www.google.com/maps/@?api=1&map_action=pano&viewpoint="
                    + lat.toFixed(6) + "," + lng.toFixed(6);
            }

            function setActive(button, isActive) {
                button.classList.toggle("street-view-picker-active", isActive);
                map.getContainer().style.cursor = isActive ? "crosshair" : "";
            }

            setTimeout(function() {
                const button = document.querySelector(".street-view-picker-button");
                if (!button) return;

                button.addEventListener("click", function(event) {
                    event.preventDefault();
                    const isActive = !button.classList.contains("street-view-picker-active");
                    setActive(button, isActive);
                });

                map.on("click", function(clickEvent) {
                    if (!button.classList.contains("street-view-picker-active")) return;
                    setActive(button, false);
                    window.open(
                        streetViewUrl(clickEvent.latlng.lat, clickEvent.latlng.lng),
                        "_blank",
                        "noopener"
                    );
                });
            }, 0);
        })();
        {% endmacro %}
        """
    )

    def __init__(self, css: str):
        super().__init__()
        self._name = "StreetViewPicker"
        self.css = json.dumps(css, ensure_ascii=False)


class InteractiveMapGenerator:
    """Генератор на интерактивна карта"""
    
    def __init__(self, config: OutputConfig):
        self.config = config
        self.run_date = datetime.now().strftime("%Y-%m-%d")
        # Зареждаме централната матрица
        self.central_matrix = get_distance_matrix_from_central_cache([])
        self.use_routing = False
        self.routing_engine = None
        
        # Определяме кой routing engine да използваме
        main_config = get_config()
        self.routing_engine = main_config.routing.engine
        
        if self.routing_engine.value == RoutingEngine.VALHALLA.value:
            # Проверяваме дали Valhalla е достъпен
            try:
                valhalla_config = main_config.valhalla
                test_url = f"{valhalla_config.base_url}/status"
                response = requests.get(test_url, timeout=5)
                if response.status_code == 200:
                    logger.info("✅ Valhalla сървър е достъпен - ще използвам реални маршрути")
                    self.use_routing = True
                else:
                    logger.warning("⚠️ Valhalla сървър не отговаря - ще използвам прави линии")
                    self.use_routing = False
            except Exception as e:
                logger.warning(f"⚠️ Не мога да се свържа с Valhalla сървъра: {e}")
                self.use_routing = False
        else:
            # Проверяваме дали OSRM е достъпен
            try:
                osrm_config = main_config.osrm
                test_url = f"{osrm_config.base_url.rstrip('/')}/route/v1/driving/23.3,42.7;23.3,42.7"
                response = requests.get(test_url, timeout=5)
                if response.status_code == 200:
                    logger.info("✅ OSRM сървър е достъпен - ще използвам реални маршрути")
                    self.use_routing = True
                else:
                    logger.warning("⚠️ OSRM сървър не отговаря - ще използвам прави линии")
                    self.use_routing = False
            except Exception as e:
                logger.warning(f"⚠️ Не мога да се свържа с OSRM сървъра: {e}")
                logger.warning("   Ще използвам прави линии за маршрутите")
                self.use_routing = False

    def _use_google_maps(self) -> bool:
        return getattr(self.config, "map_provider", "osm").strip().lower() == "google"

    def _create_folium_map(self, location: Tuple[float, float]) -> folium.Map:
        tiles = getattr(self.config, "folium_tiles", "Esri.WorldStreetMap") or "Esri.WorldStreetMap"
        custom_tiles = {
            "esri.worldstreetmap": {
                "url": "https://server.arcgisonline.com/ArcGIS/rest/services/World_Street_Map/MapServer/tile/{z}/{y}/{x}",
                "attr": "Tiles &copy; Esri",
                "name": "Esri WorldStreetMap",
            },
            "esri.worldtopomap": {
                "url": "https://server.arcgisonline.com/ArcGIS/rest/services/World_Topo_Map/MapServer/tile/{z}/{y}/{x}",
                "attr": "Tiles &copy; Esri",
                "name": "Esri WorldTopoMap",
            },
            "esri.worldimagery": {
                "url": "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
                "attr": "Tiles &copy; Esri",
                "name": "Esri WorldImagery",
            },
        }
        provider = custom_tiles.get(tiles.strip().lower())
        if provider:
            route_map = folium.Map(
                location=location,
                zoom_start=self.config.map_zoom_level,
                tiles=None,
                control_scale=True,
            )
            folium.TileLayer(
                tiles=provider["url"],
                attr=provider["attr"],
                name=provider["name"],
                overlay=False,
                control=True,
            ).add_to(route_map)
            return route_map
        return folium.Map(
            location=location,
            zoom_start=self.config.map_zoom_level,
            tiles=tiles,
            control_scale=True,
        )

    def _segment_distance_m(self, start: Tuple[float, float], end: Tuple[float, float]) -> float:
        lat1, lon1 = math.radians(start[0]), math.radians(start[1])
        lat2, lon2 = math.radians(end[0]), math.radians(end[1])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
        return 6371000 * 2 * math.asin(math.sqrt(a))

    def _bearing_degrees(self, start: Tuple[float, float], end: Tuple[float, float]) -> float:
        lat1, lon1 = math.radians(start[0]), math.radians(start[1])
        lat2, lon2 = math.radians(end[0]), math.radians(end[1])
        dlon = lon2 - lon1
        y = math.sin(dlon) * math.cos(lat2)
        x = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
        return (math.degrees(math.atan2(y, x)) + 360) % 360

    def _direction_arrow_samples(self, points: List[Tuple[float, float]], max_arrows: int = 6):
        clean_points = [point for point in points if point and len(point) == 2]
        if len(clean_points) < 2:
            return []

        segments = []
        total_m = 0.0
        for start, end in zip(clean_points, clean_points[1:]):
            distance_m = self._segment_distance_m(start, end)
            if distance_m < 5:
                continue
            segments.append((start, end, distance_m, total_m))
            total_m += distance_m

        if total_m < 25:
            return []

        arrow_count = max(1, min(max_arrows, int(total_m // 2500) + 1))
        targets = [(idx + 1) * total_m / (arrow_count + 1) for idx in range(arrow_count)]
        samples = []

        for target_m in targets:
            for start, end, distance_m, segment_start_m in segments:
                if segment_start_m <= target_m <= segment_start_m + distance_m:
                    ratio = (target_m - segment_start_m) / distance_m
                    lat = start[0] + (end[0] - start[0]) * ratio
                    lon = start[1] + (end[1] - start[1]) * ratio
                    bearing = self._bearing_degrees(start, end)
                    samples.append(((lat, lon), bearing))
                    break

        return samples

    def _add_direction_arrows(self, polyline, layer, color: str) -> None:
        """Add direction arrows along a Folium polyline."""
        try:
            if PolyLineTextPath is None or polyline is None:
                return

            marker_color = html.escape(str(color), quote=True)
            PolyLineTextPath(
                polyline,
                "        \u25ba        ",
                repeat=True,
                offset=8,
                attributes={
                    "fill": marker_color,
                    "font-weight": "bold",
                    "font-size": "14",
                    "stroke": "white",
                    "stroke-width": "2",
                    "paint-order": "stroke",
                },
            ).add_to(layer)

        except Exception as e:
            logger.debug(f"Could not add route direction arrows: {e}")

    def _get_google_maps_api_key(self) -> str:
        return (
            getattr(self.config, "google_maps_api_key", "")
            or os.environ.get("GOOGLE_MAPS_API_KEY", "")
        ).strip()

    def _json_coords(self, coords: Tuple[float, float]) -> Dict[str, float]:
        return {"lat": float(coords[0]), "lng": float(coords[1])}

    def _center_zone_map_data(self, locations) -> Optional[Dict]:
        polygon = getattr(locations, "center_zone_polygon", [])
        mode = str(getattr(locations, "center_zone_mode", "circle")).lower()
        if mode == "polygon" and len(polygon) >= 3:
            return {
                "type": "polygon",
                "path": [self._json_coords(point) for point in polygon],
            }

        return {
            "type": "circle",
            "center": self._json_coords(locations.center_location),
            "radiusMeters": locations.center_zone_radius_km * 1000,
        }

    def _html_popup(self, title: str, lines: List[str], color: str = "#222") -> str:
        body = "<br>".join(lines)
        return (
            '<div style="font-family:Arial,sans-serif;max-width:300px">'
            f'<h4 style="margin:0 0 6px;color:{html.escape(color)}">{html.escape(title)}</h4>'
            '<hr style="margin:5px 0">'
            f"{body}</div>"
        )

    def _popup_action_buttons(self, navigation_url: str, street_view_url: str) -> str:
        safe_nav = html.escape(navigation_url, quote=True)
        safe_street = html.escape(street_view_url, quote=True)
        pegman_icon = (
            '<span style="position:relative;display:inline-block;width:9px;height:13px;'
            'border-radius:5px 5px 3px 3px;background:#1f1f1f;vertical-align:-2px;'
            'margin-right:5px;">'
            '<span style="position:absolute;left:2px;top:-5px;width:5px;height:5px;'
            'border-radius:50%;background:#1f1f1f;"></span>'
            '</span>'
        )
        return (
            f'<a href="{safe_nav}" target="_blank" rel="noopener" '
            'style="display:inline-block;margin-top:8px;padding:6px 10px;'
            'background:#1a73e8;color:white;text-decoration:none;'
            'border-radius:4px;font-weight:bold;">Навигация</a>'
            f'<a href="{safe_street}" target="_blank" rel="noopener" '
            'title="Отвори Google Street View до клиента" '
            'style="display:inline-block;margin-top:8px;margin-left:6px;padding:6px 10px;'
            'background:#fbbc04;color:#1f1f1f;text-decoration:none;border-radius:4px;'
            'font-weight:bold;box-shadow:0 0 0 1px #9a6f00 inset;">'
            f'{pegman_icon}Street View</a>'
        )

    def _get_route_visual_geometry(
        self,
        route: Route,
    ) -> Tuple[List[Tuple[float, float]], bool, str]:
        route_depot = route.depot_location
        waypoints = [route_depot]
        for customer in route.customers:
            if customer.coordinates:
                waypoints.append(customer.coordinates)
        waypoints.append(route_depot)

        if not route.customers:
            return waypoints, False, "Няма клиенти"

        if not self.use_routing:
            return waypoints, True, "Прави линии"

        engine_name = (
            "Valhalla"
            if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value
            else "OSRM"
        )
        try:
            route_geometry = self._get_full_route_geometry(waypoints)
            if len(route_geometry) > 2:
                return route_geometry, False, f"{engine_name} маршрут"
        except Exception as e:
            logger.warning(f"Грешка при геометрия за Google карта: {e}")

        return waypoints, True, f"{engine_name} fallback"

    def _build_google_routes(self, routes: List[Route], start_number: int = 1) -> List[Dict[str, object]]:
        google_routes = []
        for route_idx, route in enumerate(routes):
            route_number = start_number + route_idx
            vehicle_settings = VEHICLE_SETTINGS.get(route.vehicle_type.value, {
                "color": "gray",
                "icon": "circle",
                "prefix": "fa",
                "name": "Неизвестен",
            })
            bus_color = BUS_COLORS[(route_number - 1) % len(BUS_COLORS)]
            geometry, dashed, geometry_label = self._get_route_visual_geometry(route)

            markers = []
            for client_idx, customer in enumerate(route.customers):
                if not customer.coordinates:
                    continue
                client_number = client_idx + 1
                navigation_url = self._navigation_url(customer.coordinates)
                street_view_url = self._street_view_url(customer.coordinates)
                popup_html = self._html_popup(
                    f"Автобус {route_number} - {vehicle_settings['name']}",
                    [
                        f"<b>Клиент:</b> {html.escape(str(customer.name))}",
                        f"<b>ID:</b> {html.escape(str(customer.id))}",
                        f"<b>Ред в маршрута:</b> #{client_number}",
                        f"<b>Обем:</b> {customer.volume:.2f} ст.",
                        f"<b>Координати:</b> {customer.coordinates[0]:.6f}, {customer.coordinates[1]:.6f}",
                        self._popup_action_buttons(navigation_url, street_view_url),
                    ],
                    bus_color,
                )
                markers.append({
                    "position": self._json_coords(customer.coordinates),
                    "number": client_number,
                    "title": f"#{client_number}: {customer.name}",
                    "popup": popup_html,
                    "customerName": str(customer.name),
                    "customerId": str(customer.id),
                    "volume": customer.volume,
                    "navigationUrl": navigation_url,
                    "streetViewUrl": street_view_url,
                })

            popup_html = self._html_popup(
                f"Автобус {route_number} - {vehicle_settings['name']}",
                [
                    f"<b>{geometry_label}:</b> {'прави линии' if dashed else 'реална геометрия'}",
                    f"<b>Клиенти:</b> {len(route.customers)}",
                    f"<b>Разстояние:</b> {route.total_distance_km:.1f} км",
                    f"<b>Време:</b> {route.total_time_minutes:.0f} мин",
                    f"<b>Обем:</b> {route.total_volume:.1f} ст.",
                    f"<b>Геометрия:</b> {len(geometry)} точки",
                ],
                bus_color,
            )
            google_routes.append({
                "id": f"route-{route_number}",
                "name": f"Автобус {route_number} ({len(route.customers)} клиента)",
                "color": bus_color,
                "vehicleName": vehicle_settings["name"],
                "markers": markers,
                "path": [self._json_coords(point) for point in geometry],
                "dashed": dashed,
                "popup": popup_html,
                "distanceKm": route.total_distance_km,
                "timeMin": route.total_time_minutes,
                "volume": route.total_volume,
            })
        return google_routes

    def _depot_name(self, depot: Tuple[float, float]) -> str:
        locations = get_config().locations
        if depot == locations.center_location:
            return "Център депо"
        if depot == locations.vratza_depot_location:
            return "Депо Враца"
        return "Главно депо"

    def _build_google_html(
        self,
        title: str,
        center: Tuple[float, float],
        routes: List[Route],
        depot_locations: List[Tuple[float, float]],
        single_route_number: Optional[int] = None,
    ) -> GoogleMapDocument:
        api_key = self._get_google_maps_api_key()
        if not api_key:
            return GoogleMapDocument(
                "<!doctype html><html><head><meta charset='utf-8'><title>Missing Google Maps API key</title>"
                "<style>body{font-family:Arial,sans-serif;margin:32px;line-height:1.5}</style></head><body>"
                "<h2>Missing Google Maps API key</h2>"
                "<p>Set output.google_maps_api_key in config.py or set the GOOGLE_MAPS_API_KEY environment variable.</p>"
                "</body></html>"
            )

        cfg = get_config()
        route_start = single_route_number or 1
        map_data = {
            "title": title,
            "center": self._json_coords(center),
            "zoom": self.config.map_zoom_level,
            "depots": [
                {
                    "position": self._json_coords(depot),
                    "name": self._depot_name(depot),
                    "popup": self._html_popup(self._depot_name(depot), [
                        f"<b>Координати:</b> {depot[0]:.6f}, {depot[1]:.6f}"
                    ]),
                }
                for depot in depot_locations
            ],
            "centerZone": None,
            "routes": self._build_google_routes(routes, route_start),
            "singleRouteNumber": single_route_number,
            "totals": {
                "distanceKm": sum(route.total_distance_km for route in routes),
                "timeMin": sum(route.total_time_minutes for route in routes),
                "volume": sum(route.total_volume for route in routes),
                "routeCount": len(routes),
            },
        }

        if cfg.locations.enable_center_zone_priority:
            map_data["centerZone"] = self._center_zone_map_data(cfg.locations)

        data_json = json.dumps(map_data, ensure_ascii=False)
        key_attr = html.escape(api_key, quote=True)
        html_content = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    html, body, #map {{ height: 100%; margin: 0; }}
    body {{ font-family: Arial, sans-serif; }}
    #legend {{
      position: absolute; top: 10px; left: 10px; z-index: 5; width: 290px;
      background: #fff; border: 1px solid #777; border-radius: 5px;
      box-shadow: 0 2px 12px rgba(0,0,0,.25); padding: 10px; font-size: 13px;
      max-height: calc(100vh - 40px); overflow: auto;
    }}
    #route-filter {{
      position: absolute; top: 10px; right: 10px; z-index: 5; max-width: 290px;
      background: #fff; border: 1px solid #777; border-radius: 5px;
      box-shadow: 0 2px 12px rgba(0,0,0,.25); padding: 10px; font-size: 13px;
      max-height: calc(100vh - 40px); overflow: auto;
    }}
    #route-filter.collapsed {{
      background: transparent; border: 0; box-shadow: none; padding: 0; overflow: visible;
    }}
    .client-toggle {{
      border: 1px solid rgba(35,35,35,.35); border-radius: 6px; background: #fff;
      box-shadow: 0 2px 10px rgba(0,0,0,.22); color: #222; cursor: pointer;
      font-size: 13px; font-weight: 700; padding: 8px 11px;
    }}
    .client-toggle:hover {{ background: #f4f7fb; }}
    .client-panel.hidden, .client-toggle.hidden {{ display: none; }}
    .client-panel {{
      width: 340px; max-height: calc(100vh - 90px); overflow: auto;
      background: #fff; border: 1px solid rgba(35,35,35,.35); border-radius: 6px;
      box-shadow: 0 4px 18px rgba(0,0,0,.22); padding: 10px;
    }}
    .client-panel-header {{ display: flex; align-items: center; justify-content: space-between; gap: 8px; }}
    .client-panel-header b {{ font-size: 15px; }}
    .client-close {{
      width: 26px; height: 26px; border: 0; border-radius: 4px; background: #f1f3f4;
      color: #333; cursor: pointer; font-size: 18px; line-height: 1; font-weight: 700;
    }}
    .client-close:hover {{ background: #e4e7eb; }}
    .client-route-stats {{
      display: grid; grid-template-columns: 1fr 1fr; gap: 6px;
      margin: 8px 0 10px; padding: 8px; background: #f8fafc;
      border: 1px solid #e1e6ee; border-radius: 6px;
    }}
    .client-route-stat {{
      display: flex; justify-content: space-between; gap: 8px;
      font-size: 12px; color: #1f2933;
    }}
    .client-route-stat b {{ font-size: 12px; }}
    .route-row {{ display: flex; align-items: center; gap: 6px; margin: 4px 0; }}
    .swatch {{ width: 14px; height: 14px; display: inline-block; border-radius: 2px; }}
    .client-actions {{ display: flex; gap: 5px; flex-wrap: wrap; justify-content: flex-end; }}
    .client-action {{
      display: inline-flex; align-items: center; justify-content: center; gap: 4px;
      min-height: 24px; padding: 5px 7px; border-radius: 4px;
      text-decoration: none; font-weight: 700; font-size: 11px; line-height: 1;
    }}
    .client-nav {{ background: #1a73e8; color: #fff; }}
    .client-street {{ background: #fbbc04; color: #1f1f1f; box-shadow: 0 0 0 1px #9a6f00 inset; }}
    .pegman-icon {{
      position: relative; display: inline-block; width: 9px; height: 13px;
      border-radius: 5px 5px 3px 3px; background: #1f1f1f; flex: 0 0 auto;
    }}
    .pegman-icon::before {{
      content: ""; position: absolute; left: 2px; top: -5px; width: 5px; height: 5px;
      border-radius: 50%; background: #1f1f1f;
    }}
    .marker-label {{
      color: #fff; font-weight: 700; font-size: 13px; text-align: center;
      text-shadow: 0 1px 2px rgba(0,0,0,.8);
    }}
  </style>
</head>
<body>
  <div id="map"></div>
  <div id="legend"></div>
  <div id="route-filter"></div>
  <script>
    const MAP_DATA = {data_json};
    let infoWindow;

    function initMap() {{
      const map = new google.maps.Map(document.getElementById("map"), {{
        center: MAP_DATA.center,
        zoom: MAP_DATA.zoom,
        mapTypeControl: true,
        streetViewControl: true,
        fullscreenControl: true
      }});
      infoWindow = new google.maps.InfoWindow();
      const bounds = new google.maps.LatLngBounds();
      const routeObjects = {{}};

      MAP_DATA.depots.forEach((depot) => {{
        const marker = new google.maps.Marker({{
          position: depot.position,
          map,
          title: depot.name,
          label: {{ text: "D", color: "white", fontWeight: "bold" }},
          icon: {{
            path: google.maps.SymbolPath.CIRCLE,
            scale: 12,
            fillColor: "#111111",
            fillOpacity: 1,
            strokeColor: "#ffffff",
            strokeWeight: 2
          }}
        }});
        marker.addListener("click", () => infoWindow.setContent(depot.popup) || infoWindow.open(map, marker));
        bounds.extend(depot.position);
      }});

      if (MAP_DATA.centerZone) {{
        if (MAP_DATA.centerZone.type === "polygon") {{
          new google.maps.Polygon({{
            map,
            paths: MAP_DATA.centerZone.path,
            strokeColor: "#d32f2f",
            strokeOpacity: 0.9,
            strokeWeight: 2,
            fillColor: "#d32f2f",
            fillOpacity: 0.12
          }});
          MAP_DATA.centerZone.path.forEach((point) => bounds.extend(point));
        }} else {{
          new google.maps.Circle({{
            map,
            center: MAP_DATA.centerZone.center,
            radius: MAP_DATA.centerZone.radiusMeters,
            strokeColor: "#d32f2f",
            strokeOpacity: 0.9,
            strokeWeight: 2,
            fillColor: "#d32f2f",
            fillOpacity: 0.12
          }});
          bounds.extend(MAP_DATA.centerZone.center);
        }}
      }}

      MAP_DATA.routes.forEach((route) => {{
        const directionArrow = {{
          path: google.maps.SymbolPath.FORWARD_CLOSED_ARROW,
          scale: 2.8,
          strokeColor: "#ffffff",
          strokeOpacity: 0.95,
          strokeWeight: 2,
          fillColor: route.color,
          fillOpacity: 0.95
        }};
        const directionIcons = ["35%", "70%"].map((offset) => ({{
          icon: directionArrow,
          offset
        }}));
        const polylineOptions = {{
          path: route.path,
          map,
          strokeColor: route.color,
          strokeOpacity: route.dashed ? 0 : 0.82,
          strokeWeight: route.dashed ? 0 : 4
        }};
        if (route.dashed) {{
          polylineOptions.icons = [{{
            icon: {{ path: "M 0,-1 0,1", strokeOpacity: 0.85, scale: 3, strokeColor: route.color }},
            offset: "0",
            repeat: "14px"
          }}, ...directionIcons];
        }} else {{
          polylineOptions.icons = directionIcons;
        }}
        const line = new google.maps.Polyline(polylineOptions);
        line.addListener("click", (event) => {{
          infoWindow.setContent(route.popup);
          infoWindow.setPosition(event.latLng);
          infoWindow.open(map);
        }});

        const markerObjects = route.markers.map((point) => {{
          const marker = new google.maps.Marker({{
            position: point.position,
            map,
            title: point.title,
            label: {{ text: String(point.number), color: "white", fontWeight: "bold" }},
            icon: {{
              path: google.maps.SymbolPath.CIRCLE,
              scale: 13,
              fillColor: route.color,
              fillOpacity: 1,
              strokeColor: "#ffffff",
              strokeWeight: 3
            }}
          }});
          marker.addListener("click", () => infoWindow.setContent(point.popup) || infoWindow.open(map, marker));
          bounds.extend(point.position);
          return {{ marker, data: point }};
        }});

        route.path.forEach((point) => bounds.extend(point));
        routeObjects[route.id] = {{
          line,
          markers: markerObjects.map((item) => item.marker),
          markerObjects
        }};
      }});

      if (!bounds.isEmpty()) {{
        map.fitBounds(bounds, 40);
      }}
      renderLegend();
      renderFilter(routeObjects, map);
    }}

    function renderLegend() {{
      const totals = MAP_DATA.totals;
      document.getElementById("legend").innerHTML = `
        <h4 style="margin:0 0 8px;text-align:center">${{MAP_DATA.title}}</h4>
        <div><b>Депо:</b> черен маркер D</div>
        <hr>
        <div><b>Статистики:</b></div>
        <div>Общо разстояние: ${{totals.distanceKm.toFixed(1)}} км</div>
        <div>Общо време: ${{totals.timeMin.toFixed(0)}} мин</div>
        <div>Общ обем: ${{totals.volume.toFixed(1)}} ст.</div>
        <div>Маршрути: ${{totals.routeCount}}</div>
      `;
    }}

    function renderFilter(routeObjects, map) {{
      const container = document.getElementById("route-filter");
      if (MAP_DATA.singleRouteNumber) {{
        renderCustomerPanel(container, routeObjects, map);
        return;
      }}
      container.innerHTML = "<b>Филтър на автобуси</b>";
      MAP_DATA.routes.forEach((route) => {{
        const row = document.createElement("label");
        row.className = "route-row";
        row.innerHTML = `
          <input type="checkbox" checked data-route="${{route.id}}">
          <span class="swatch" style="background:${{route.color}}"></span>
          <span>${{route.name}}</span>
        `;
        container.appendChild(row);
      }});
      Object.values(routeObjects).forEach((objects) => {{
        const map = objects.line.getMap();
        objects.line.set("mapRef", map);
      }});
      container.querySelectorAll("input[type=checkbox]").forEach((checkbox) => {{
        checkbox.addEventListener("change", () => {{
          const objects = routeObjects[checkbox.dataset.route];
          const map = checkbox.checked ? objects.line.get("mapRef") : null;
          objects.line.setMap(map);
          objects.markers.forEach((marker) => marker.setMap(map));
        }});
      }});
    }}

    function formatRouteDuration(minutes) {{
      const total = Math.max(0, Math.round(Number(minutes) || 0));
      const hours = Math.floor(total / 60);
      const mins = total % 60;
      if (hours > 0 && mins > 0) return `${{hours}}ч ${{mins}}м`;
      if (hours > 0) return `${{hours}}ч`;
      return `${{mins}}м`;
    }}

    function formatRouteNumber(value) {{
      const number = Number(value) || 0;
      return Math.abs(number - Math.round(number)) < 0.05
        ? String(Math.round(number))
        : number.toFixed(1);
    }}

    function renderCustomerPanel(container, routeObjects, map) {{
      const route = MAP_DATA.routes[0];
      const objects = routeObjects[route.id];
      const routeTimeText = formatRouteDuration(route.timeMin);
      container.classList.add("collapsed");
      container.innerHTML = `
        <button type="button" class="client-toggle">Клиенти</button>
        <div class="client-panel hidden">
          <div class="client-panel-header">
            <b>Клиенти - маршрут ${{MAP_DATA.singleRouteNumber}}</b>
            <button type="button" class="client-close" title="Затвори">×</button>
          </div>
          <div class="client-route-stats">
            <div class="client-route-stat"><span>Общо:</span><b>${{route.markers.length}} клиента</b></div>
            <div class="client-route-stat"><span>Стекове:</span><b>${{formatRouteNumber(route.volume)}}</b></div>
            <div class="client-route-stat"><span>Време:</span><b>${{routeTimeText}}</b></div>
            <div class="client-route-stat"><span>Км:</span><b>${{formatRouteNumber(route.distanceKm)}}</b></div>
          </div>
          <div class="client-list"></div>
        </div>
      `;
      const toggle = container.querySelector(".client-toggle");
      const panel = container.querySelector(".client-panel");
      const close = container.querySelector(".client-close");
      const list = container.querySelector(".client-list");
      function setOpen(isOpen) {{
        container.classList.toggle("collapsed", !isOpen);
        panel.classList.toggle("hidden", !isOpen);
        toggle.classList.toggle("hidden", isOpen);
      }}
      toggle.addEventListener("click", () => setOpen(true));
      close.addEventListener("click", () => setOpen(false));
      route.markers.forEach((point, index) => {{
        const row = document.createElement("div");
        row.className = "route-row";
        row.style.cursor = "pointer";
        row.style.borderTop = "1px solid #ececec";
        row.style.padding = "7px 0";
        row.innerHTML = `
          <span class="swatch" style="background:${{route.color}};border-radius:50%;color:#fff;text-align:center;line-height:14px;font-size:10px;font-weight:bold">${{point.number}}</span>
          <span style="flex:1;min-width:0">
            <span style="display:block;font-weight:bold;font-size:12px;overflow-wrap:anywhere">${{point.customerName}}</span>
            <span style="display:block;color:#666;font-size:11px">ID: ${{point.customerId}} · ${{Number(point.volume).toFixed(2)}} ст.</span>
          </span>
          <span class="client-actions">
            <a class="client-action client-nav" href="${{point.navigationUrl}}" target="_blank" rel="noopener"
               title="Отвори навигация до клиента">Навигация</a>
            <a class="client-action client-street" href="${{point.streetViewUrl}}" target="_blank" rel="noopener"
               title="Отвори Google Street View до клиента">
              <span class="pegman-icon" aria-hidden="true"></span>
              <span>Street View</span>
            </a>
          </span>
        `;
        row.addEventListener("click", (event) => {{
          if (event.target.closest("a")) return;
          const marker = objects.markerObjects[index].marker;
          map.panTo(marker.getPosition());
          map.setZoom(Math.max(map.getZoom(), 15));
          infoWindow.setContent(point.popup);
          infoWindow.open(map, marker);
        }});
        list.appendChild(row);
      }});
    }}
  </script>
  <script async defer src="https://maps.googleapis.com/maps/api/js?key={key_attr}&callback=initMap"></script>
</body>
</html>
"""
        return GoogleMapDocument(html_content)
    
    def create_map(self, solution: CVRPSolution, warehouse_allocation: WarehouseAllocation,
                  depot_location: Tuple[float, float]) -> folium.Map:
        """Създава интерактивна карта с маршрутите"""
        logger.info("Създавам интерактивна карта")
        
        # Показваме routing статуса
        if self.use_routing:
            engine_name = "Valhalla" if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value else "OSRM"
            logger.info(f"🛣️ Използвам {engine_name} Route API за реални маршрути")
        else:
            logger.warning("📐 Използвам прави линии (Routing engine недостъпен)")
            
        # Взимаме всички уникални депа от маршрутите
        unique_depots = {depot_location}  # Добавяме основното депо
        
        # Добавяме депата от маршрутите
        for route in solution.routes:
            if hasattr(route, 'depot_location') and route.depot_location:
                unique_depots.add(route.depot_location)

        if self._use_google_maps():
            return self._build_google_html(
                "Интерактивна карта",
                depot_location,
                solution.routes,
                list(unique_depots),
            )
        
        # Инициализация на картата с основното депо като център
        route_map = self._create_folium_map(depot_location)
        
        # Добавяне на всички депа
        self._add_depot_markers(route_map, list(unique_depots))
        
        # Добавяне на център зоната
        from config import get_config
        locations = get_config().locations
        if locations.enable_center_zone_priority:
            self._add_center_zone_shape(route_map, locations)
        
        # Добавяне на маршрутите с OSRM геометрия
        if self.config.show_route_colors:
            self._add_routes_to_map(route_map, solution.routes)
        
        # Добавяне на легенда
        self._add_legend(route_map, solution.routes)
        
        return route_map
    
    def _add_depot_markers(self, route_map: folium.Map, depot_locations: List[Tuple[float, float]]):
        """Добавя маркери за всички депа"""
        from config import get_config
        locations = get_config().locations
        
        for i, depot in enumerate(depot_locations):
            # Определяме кое депо е това
            depot_name = "Главно депо"
            if depot == locations.center_location:
                depot_name = "Център депо"
            elif depot == locations.vratza_depot_location:
                depot_name = "Депо Враца"
            
            # Добавяме специален маркер за всяко депо
            folium.Marker(
                depot,
                popup=f"<b>{depot_name}</b>",
                tooltip=depot_name,
                icon=folium.Icon(color='black', icon='home', prefix='fa')
            ).add_to(route_map)
            
    def _add_depot_marker(self, route_map: folium.Map, depot_location: Tuple[float, float]):
        """Добавя маркер за едно депо (поддържа се за обратна съвместимост)"""
        self._add_depot_markers(route_map, [depot_location])
    
    def _add_center_zone_circle(self, route_map: folium.Map, center_location: Tuple[float, float], radius_km: float):
        """Добавя кръг за център зоната"""
        folium.Circle(
            location=center_location,
            radius=radius_km * 1000,  # Конвертираме в метри
            color='red',
            fill=True,
            fillColor='red',
            fillOpacity=0.1,
            popup=f"<b>Център зона</b><br>Радиус: {radius_km} км",
            tooltip="Център зона"
        ).add_to(route_map)
        
        # Добавяме маркер за центъра
        folium.Marker(
            location=center_location,
            popup=f"<b>Център</b><br>Координати: {center_location[0]:.6f}, {center_location[1]:.6f}<br>Радиус зона: {radius_km} км",
            icon=folium.Icon(color='red', icon='star'),
            tooltip="Център"
        ).add_to(route_map)

    def _add_center_zone_shape(self, route_map: folium.Map, locations):
        polygon = getattr(locations, "center_zone_polygon", [])
        mode = str(getattr(locations, "center_zone_mode", "circle")).lower()
        if mode == "polygon" and len(polygon) >= 3:
            folium.Polygon(
                locations=polygon,
                color="red",
                fill=True,
                fillColor="red",
                fillOpacity=0.1,
                popup=f"<b>Център зона</b><br>Полигон: {len(polygon)} точки",
                tooltip="Център зона",
            ).add_to(route_map)
            folium.Marker(
                location=locations.center_location,
                popup=f"<b>Център</b><br>Координати: {locations.center_location[0]:.6f}, {locations.center_location[1]:.6f}",
                icon=folium.Icon(color="red", icon="star"),
                tooltip="Център",
            ).add_to(route_map)
            return

        self._add_center_zone_circle(route_map, locations.center_location, locations.center_zone_radius_km)
    
    def _get_osrm_route_geometry(self, start_coords: Tuple[float, float],
                                end_coords: Tuple[float, float]) -> List[Tuple[float, float]]:
        """Получава реална геометрия на маршрута от OSRM Route API"""
        try:
            import requests
            from config import get_config
            
            # OSRM Route API заявка за пълна геометрия
            osrm_config = get_config().osrm
            base_url = osrm_config.base_url.rstrip('/')
            
            # Форматираме координатите за OSRM (lon,lat формат)
            start_lon, start_lat = start_coords[1], start_coords[0]
            end_lon, end_lat = end_coords[1], end_coords[0]
            
            route_url = f"{base_url}/route/v1/driving/{start_lon:.6f},{start_lat:.6f};{end_lon:.6f},{end_lat:.6f}?geometries=geojson&overview=full&steps=false"
            
            response = requests.get(route_url, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            if data['code'] == 'Ok' and data['routes']:
                route = data['routes'][0]
                coordinates = route['geometry']['coordinates']
                
                # Конвертираме от [lon,lat] към [lat,lon] за Folium
                geometry = [(coord[1], coord[0]) for coord in coordinates]
                
                logger.debug(f"✅ OSRM геометрия получена: {len(geometry)} точки")
                return geometry
            else:
                logger.warning(f"OSRM Route API грешка: {data.get('message', 'Неизвестна грешка')}")
            return [start_coords, end_coords]
            
        except Exception as e:
            logger.warning(f"Грешка при OSRM Route API заявка: {e}")
            # Fallback към права линия
            return [start_coords, end_coords]
    
    def _get_valhalla_route_geometry(self, start_coords: Tuple[float, float],
                                     end_coords: Tuple[float, float]) -> List[Tuple[float, float]]:
        """Получава реална геометрия на маршрута от Valhalla Route API"""
        try:
            import requests
            import json
            from config import get_config
            
            valhalla_config = get_config().valhalla
            base_url = valhalla_config.base_url.rstrip('/')
            
            # Valhalla Route API заявка
            route_request = {
                "locations": [
                    {"lat": start_coords[0], "lon": start_coords[1]},
                    {"lat": end_coords[0], "lon": end_coords[1]}
                ],
                "costing": valhalla_config.costing,
                "directions_options": {
                    "units": "kilometers"
                }
            }
            
            route_url = f"{base_url}/route"
            response = requests.post(route_url, json=route_request, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            if 'trip' in data and 'legs' in data['trip']:
                geometry = []
                for leg in data['trip']['legs']:
                    if 'shape' in leg:
                        # Valhalla връща encoded polyline, трябва да го декодираме
                        decoded = self._decode_polyline(leg['shape'])
                        geometry.extend(decoded)
                
                if geometry:
                    logger.debug(f"✅ Valhalla геометрия получена: {len(geometry)} точки")
                    return geometry
            
            logger.warning(f"Valhalla Route API грешка: Няма геометрия в отговора")
            return [start_coords, end_coords]
            
        except Exception as e:
            logger.warning(f"Грешка при Valhalla Route API заявка: {e}")
            return [start_coords, end_coords]
    
    def _decode_polyline(self, encoded: str, precision: int = 6) -> List[Tuple[float, float]]:
        """Декодира Google encoded polyline (използва се от Valhalla)"""
        inv = 1.0 / (10 ** precision)
        decoded = []
        previous = [0, 0]
        i = 0
        
        while i < len(encoded):
            ll = [0, 0]
            for j in range(2):
                shift = 0
                byte = 0x20
                
                while byte >= 0x20:
                    byte = ord(encoded[i]) - 63
                    i += 1
                    ll[j] |= (byte & 0x1f) << shift
                    shift += 5
                
                ll[j] = previous[j] + (~(ll[j] >> 1) if ll[j] & 1 else (ll[j] >> 1))
                previous[j] = ll[j]
            
            decoded.append((ll[0] * inv, ll[1] * inv))
        
        return decoded
    
    def _get_route_geometry(self, start_coords: Tuple[float, float],
                           end_coords: Tuple[float, float]) -> List[Tuple[float, float]]:
        """Получава геометрия на маршрута от избрания routing engine"""
        if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value:
            return self._get_valhalla_route_geometry(start_coords, end_coords)
        else:
            return self._get_osrm_route_geometry(start_coords, end_coords)
    
    def _get_full_route_geometry(self, waypoints: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
        """Получава пълната геометрия за маршрут с множество точки.
        Ако има твърде много точки, използваме fallback за по-бърза работа.
        """
        if len(waypoints) < 2:
            return waypoints

        # ОПТИМИЗАЦИЯ: Ако маршрутът има твърде много точки, не търсим пълна геометрия,
        # а чертаем сегменти, за да не претоварваме routing engine и да ускорим процеса.
        MAX_WAYPOINTS_FOR_FULL_GEOMETRY = 50
        if len(waypoints) > MAX_WAYPOINTS_FOR_FULL_GEOMETRY:
            logger.info(f"🌀 Маршрутът има {len(waypoints)} точки (> {MAX_WAYPOINTS_FOR_FULL_GEOMETRY}). "
                        f"Използвам опростена геометрия (сегменти) за по-бърза работа.")
            full_geometry = []
            for i in range(len(waypoints) - 1):
                # За всеки сегмент взимаме геометрията (или права линия при грешка)
                segment_geometry = self._get_route_geometry(waypoints[i], waypoints[i+1])
                if i > 0:
                    # Премахваме първата точка, за да няма дублиране
                    segment_geometry = segment_geometry[1:]
                full_geometry.extend(segment_geometry)
            return full_geometry

        # Използваме подходящия routing engine
        if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value:
            return self._get_full_route_geometry_valhalla(waypoints)
        else:
            return self._get_full_route_geometry_osrm(waypoints)
    
    def _get_full_route_geometry_valhalla(self, waypoints: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
        """Получава пълната геометрия за маршрут с множество точки от Valhalla."""
        try:
            import requests
            import json
            from config import get_config
            
            valhalla_config = get_config().valhalla
            base_url = valhalla_config.base_url.rstrip('/')
            
            # Форматираме locations за Valhalla
            locations = [{"lat": lat, "lon": lon} for lat, lon in waypoints]
            
            route_request = {
                "locations": locations,
                "costing": valhalla_config.costing,
                "directions_options": {
                    "units": "kilometers"
                }
            }
            
            route_url = f"{base_url}/route"
            response = requests.post(route_url, json=route_request, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            
            if 'trip' in data and 'legs' in data['trip']:
                geometry = []
                for leg in data['trip']['legs']:
                    if 'shape' in leg:
                        decoded = self._decode_polyline(leg['shape'])
                        if geometry:
                            # Премахваме първата точка, за да няма дублиране
                            decoded = decoded[1:] if decoded else decoded
                        geometry.extend(decoded)
                
                if geometry:
                    logger.info(f"✅ Valhalla маршрут геометрия получена: {len(geometry)} точки за {len(waypoints)} waypoints")
                    return geometry
            
            logger.warning(f"Valhalla Route API грешка за пълен маршрут: Няма геометрия")
            return waypoints
            
        except Exception as e:
            logger.warning(f"Грешка при Valhalla Route API заявка за пълен маршрут: {e}")
            # Fallback към последователност от сегменти
            full_geometry = []
            for i in range(len(waypoints) - 1):
                segment = self._get_valhalla_route_geometry(waypoints[i], waypoints[i + 1])
                if i == 0:
                    full_geometry.extend(segment)
                else:
                    full_geometry.extend(segment[1:])
            
            return full_geometry if full_geometry else waypoints
    
    def _get_full_route_geometry_osrm(self, waypoints: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
        """Получава пълната геометрия за маршрут с множество точки от OSRM."""
        try:
            import requests
            from config import get_config
            
            # OSRM Route API заявка за целия маршрут
            osrm_config = get_config().osrm
            base_url = osrm_config.base_url.rstrip('/')
            
            # Форматираме всички координати за OSRM (lon,lat формат)
            coords_str = ';'.join([f"{lon:.6f},{lat:.6f}" for lat, lon in waypoints])
            
            route_url = f"{base_url}/route/v1/driving/{coords_str}?geometries=geojson&overview=full&steps=false"
            
            response = requests.get(route_url, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            
            if data['code'] == 'Ok' and data['routes']:
                route = data['routes'][0]
                coordinates = route['geometry']['coordinates']
                
                # Конвертираме от [lon,lat] към [lat,lon] за Folium
                geometry = [(coord[1], coord[0]) for coord in coordinates]
                
                logger.info(f"✅ OSRM маршрут геометрия получена: {len(geometry)} точки за {len(waypoints)} waypoints")
                return geometry
            else:
                logger.warning(f"OSRM Route API грешка за пълен маршрут: {data.get('message', 'Неизвестна грешка')}")
                return waypoints
                
        except Exception as e:
            logger.warning(f"Грешка при OSRM Route API заявка за пълен маршрут: {e}")
            # Fallback към последователност от прави линии
            full_geometry = []
            for i in range(len(waypoints) - 1):
                segment = self._get_osrm_route_geometry(waypoints[i], waypoints[i + 1])
                if i == 0:
                    full_geometry.extend(segment)
                else:
                    full_geometry.extend(segment[1:])  # Пропускаме дублираната точка
            
            return full_geometry if full_geometry else waypoints
    
    def _add_routes_to_map(self, route_map: folium.Map, routes: List[Route]):
        """Добавя маршрутите на картата с OSRM геометрия и филтър за бусовете"""
        # Създаваме FeatureGroup за всеки автобус
        bus_layers = {}
        
        for route_idx, route in enumerate(routes):
            vehicle_settings = VEHICLE_SETTINGS.get(route.vehicle_type.value, {
                'color': 'gray', 
                'icon': 'circle',
                'prefix': 'fa',
                'name': 'Неизвестен'
            })
            
            # Всеки автобус получава уникален цвят
            bus_color = BUS_COLORS[route_idx % len(BUS_COLORS)]
            bus_id = f"bus_{route_idx + 1}"
            
            # Създаваме FeatureGroup за този автобус
            bus_layer = folium.FeatureGroup(name=f"🚌 Автобус {route_idx + 1} ({len(route.customers)} клиента)")
            bus_layers[bus_id] = bus_layer
            
            # Добавяне на клиентските маркери с номерация
            for client_idx, customer in enumerate(route.customers):
                if customer.coordinates:
                    # Създаваме номериран маркер
                    client_number = client_idx + 1
                    navigation_url = self._navigation_url(customer.coordinates)
                    street_view_url = self._street_view_url(customer.coordinates)
                    
                    # HTML за номерирано пинче с уникален цвят на автобуса
                    icon_html = f'''
                    <div style="
                        background-color: {bus_color};
                        border: 3px solid white;
                        border-radius: 50%;
                        width: 30px;
                        height: 30px;
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        font-weight: bold;
                        font-size: 14px;
                        color: white;
                        text-shadow: 1px 1px 1px rgba(0,0,0,0.7);
                    ">{client_number}</div>
                    '''
                    
                    popup_text = f"""
                    <div style="font-family: Arial, sans-serif;">
                        <h4 style="margin: 0; color: {bus_color};">
                            Автобус {route_idx + 1} - {vehicle_settings['name']}
                        </h4>
                        <hr style="margin: 5px 0;">
                        <b>Клиент:</b> {customer.name}<br>
                        <b>ID:</b> {customer.id}<br>
                        <b>Ред в маршрута:</b> #{client_number}<br>
                        <b>Обем:</b> {customer.volume:.2f} ст.<br>
                        <b>Координати:</b> {customer.coordinates[0]:.6f}, {customer.coordinates[1]:.6f}<br>
                        {self._popup_action_buttons(navigation_url, street_view_url)}
                    </div>
                    """
                    
                    # Добавяме номерирания маркер в слоя на автобуса
                    marker = folium.Marker(
                        customer.coordinates,
                        popup=folium.Popup(popup_text, max_width=300),
                        tooltip=f"#{client_number}: {customer.name}",
                        icon=folium.DivIcon(
                            html=icon_html,
                            icon_size=(30, 30),
                            icon_anchor=(15, 15),
                            popup_anchor=(0, -15)
                        )
                    )
                    marker.add_to(bus_layer)
            
            # Създаваме пълния маршрут: депо -> клиенти -> депо
            if route.customers and self.use_routing:
                engine_name = "Valhalla" if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value else "OSRM"
                logger.info(f"🛣️ Получавам {engine_name} маршрут за Автобус {route_idx + 1} с {len(route.customers)} клиента")
                
                # Използваме depot_location от самия маршрут
                route_depot = route.depot_location
                
                # Подготвяме всички waypoints
                waypoints = [route_depot]
                for customer in route.customers:
                    if customer.coordinates:
                        waypoints.append(customer.coordinates)
                waypoints.append(route_depot)  # Връщане в депото
                
                # Определяме името на routing engine
                engine_name = "Valhalla" if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value else "OSRM"
                
                # Получаваме реалната геометрия
                try:
                    route_geometry = self._get_full_route_geometry(waypoints)
                    
                    if len(route_geometry) > 2:
                        # Създаваме popup с информация за маршрута
                        popup_text = f"""
                        <div style="font-family: Arial, sans-serif;">
                            <h4 style="margin: 0; color: {bus_color};">
                                🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                            </h4>
                            <hr style="margin: 5px 0;">
                            <b>{engine_name} маршрут:</b> ✅<br>
                            <b>Клиенти:</b> {len(route.customers)}<br>
                            <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                            <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                            <b>Обем:</b> {route.total_volume:.1f} ст.<br>
                            <b>Геометрия:</b> {len(route_geometry)} точки
                        </div>
                        """
                        
                        # Създаваме линията в слоя на автобуса
                        polyline = folium.PolyLine(
                            route_geometry,
                            color=bus_color,
                            weight=4,
                            opacity=0.8,
                            popup=folium.Popup(popup_text, max_width=300)
                        )
                        polyline.add_to(bus_layer)
                        self._add_direction_arrows(polyline, bus_layer, bus_color)
                        logger.info(f"✅ {engine_name} маршрут добавен за Автобус {route_idx + 1}: {len(route_geometry)} точки")
                    else:
                        # Fallback към прави линии
                        popup_text = f"""
                        <div style="font-family: Arial, sans-serif;">
                            <h4 style="margin: 0; color: {bus_color};">
                                🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                            </h4>
                            <hr style="margin: 5px 0;">
                            <b>{engine_name} маршрут:</b> ⚠️ (прави линии)<br>
                            <b>Клиенти:</b> {len(route.customers)}<br>
                            <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                            <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                            <b>Обем:</b> {route.total_volume:.1f} ст.
                        </div>
                        """
                        
                        polyline = folium.PolyLine(
                            waypoints,
                            color=bus_color,
                            weight=3,
                            opacity=0.6,
                            popup=folium.Popup(popup_text, max_width=300),
                            dashArray='5, 5'  # Пунктирана линия за показване че не е реална геометрия
                        )
                        polyline.add_to(bus_layer)
                        self._add_direction_arrows(polyline, bus_layer, bus_color)
                        logger.warning(f"⚠️ Използвам прави линии за Автобус {route_idx + 1}")
                        
                except Exception as e:
                    logger.error(f"❌ Грешка при {engine_name} маршрут за Автобус {route_idx + 1}: {e}")
                    # Fallback към прави линии
                    route_depot = route.depot_location
                    waypoints = [route_depot]
                    for customer in route.customers:
                        if customer.coordinates:
                            waypoints.append(customer.coordinates)
                    waypoints.append(route_depot)
                    
                    popup_text = f"""
                    <div style="font-family: Arial, sans-serif;">
                        <h4 style="margin: 0; color: {bus_color};">
                            🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                        </h4>
                        <hr style="margin: 5px 0;">
                        <b>{engine_name} маршрут:</b> ❌ (fallback)<br>
                        <b>Клиенти:</b> {len(route.customers)}<br>
                        <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                        <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                        <b>Обем:</b> {route.total_volume:.1f} ст.
                    </div>
                    """
                    
                    polyline = folium.PolyLine(
                        waypoints,
                        color=bus_color,
                        weight=3,
                        opacity=0.6,
                        popup=folium.Popup(popup_text, max_width=300),
                        dashArray='5, 5'
                    )
                    polyline.add_to(bus_layer)
                    self._add_direction_arrows(polyline, bus_layer, bus_color)
            
            elif route.customers:
                # Fallback към прави линии ако routing е изключен
                route_depot = route.depot_location
                waypoints = [route_depot]
                for customer in route.customers:
                    if customer.coordinates:
                        waypoints.append(customer.coordinates)
                waypoints.append(route_depot)
                
                polyline = folium.PolyLine(
                    waypoints,
                    color=bus_color,
                    weight=3,
                    opacity=0.8,
                    popup=f"🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}"
                )
                polyline.add_to(bus_layer)
                self._add_direction_arrows(polyline, bus_layer, bus_color)
        
        # Добавяме всички слоеве на автобусите към картата
        for bus_layer in bus_layers.values():
            bus_layer.add_to(route_map)
        
        # Добавяме LayerControl за филтър
        folium.LayerControl(
            position='topright',
            collapsed=True,
            overlay=True,
            control=True
                ).add_to(route_map)
    
    def _add_legend(self, route_map: folium.Map, routes: List[Route]):
        """Добавя легенда на картата с информация за маршрутите"""
        # Изчисляваме статистики
        total_distance = sum(route.total_distance_km for route in routes)
        total_time = sum(route.total_time_minutes for route in routes)
        total_volume = sum(route.total_volume for route in routes)
        routed_count = sum(1 for route in routes if self.use_routing)
        
        legend_html = f'''
        <div style="position: fixed; 
                    top: 10px; left: 10px; width: 280px; height: auto; 
                    background-color: white; border:2px solid grey; z-index:9999; 
                    font-size:14px; padding: 10px; border-radius: 5px;
                    box-shadow: 0 0 15px rgba(0,0,0,0.2);">
        <h4 style="margin-top:0; margin-bottom:10px; text-align: center;">🗺️ Информация</h4>
        '''
        
        # Добавяме депо
        legend_html += '''
        <p style="margin: 5px 0;">
            <i class="fa fa-home" style="color: black; margin-right: 8px;"></i>
            Депо
        </p>
        '''
        
        # Добавяме информация за филтъра
        legend_html += '''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-weight: bold;">🚌 Филтър на автобуси:</p>
        <p style="margin: 5px 0; font-size: 12px; color: #666;">
            Използвай контрола в горния десен ъгъл за показване/скриване на отделни автобуси
            </p>
            '''
        
        # Добавяме информация за маршрутите
        if self.use_routing:
            engine_name = "Valhalla" if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value else "OSRM"
            routing_status = f"🛣️ {engine_name} маршрути"
        else:
            routing_status = "📐 Прави линии"
        
        legend_html += f'''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-size: 12px; color: #666;">
            Числата показват реда на посещение<br>
            {routing_status}
        </p>
        '''
        
        # Добавяме статистики
        legend_html += f'''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-size: 12px; font-weight: bold;">
            📊 Статистики:
        </p>
        <p style="margin: 3px 0; font-size: 11px; color: #555;">
            • Общо разстояние: {total_distance:.1f} км<br>
            • Общо време: {total_time:.0f} мин<br>
            • Общ обем: {total_volume:.1f} ст.<br>
            • Маршрути с геометрия: {routed_count}/{len(routes)}
        </p>
        </div>
        '''
        
        # Добавяме легендата към картата
        legend_element = folium.Element(legend_html)
        route_map.get_root().add_child(legend_element)

    def _navigation_url(self, coords: Tuple[float, float]) -> str:
        return (
            "https://www.google.com/maps/dir/?api=1&destination="
            f"{coords[0]:.6f},{coords[1]:.6f}&travelmode=driving"
        )

    def _format_route_duration(self, minutes: float) -> str:
        total_minutes = max(0, int(round(minutes or 0)))
        hours, mins = divmod(total_minutes, 60)
        if hours and mins:
            return f"{hours}ч {mins}м"
        if hours:
            return f"{hours}ч"
        return f"{mins}м"

    def _format_route_number(self, value: float) -> str:
        number = float(value or 0)
        if abs(number - round(number)) < 0.05:
            return str(int(round(number)))
        return f"{number:.1f}"

    def _route_stats_html(self, route: Route) -> str:
        return f'''
            <div class="route-client-stats">
                <div class="route-client-stat"><span>Общо:</span><b>{len(route.customers)} клиента</b></div>
                <div class="route-client-stat"><span>Стекове:</span><b>{self._format_route_number(route.total_volume)}</b></div>
                <div class="route-client-stat"><span>Време:</span><b>{self._format_route_duration(route.total_time_minutes)}</b></div>
                <div class="route-client-stat"><span>Км:</span><b>{self._format_route_number(route.total_distance_km)}</b></div>
            </div>
        '''

    def _street_view_url(self, coords: Tuple[float, float]) -> str:
        return (
            "https://www.google.com/maps/@?api=1&map_action=pano&viewpoint="
            f"{coords[0]:.6f},{coords[1]:.6f}"
        )

    def _street_view_picker_css(self) -> str:
        return """
            .street-view-picker-control {
                font-family: Arial, sans-serif;
            }
            .street-view-picker-button {
                display: flex;
                align-items: center;
                gap: 7px;
                border: 1px solid rgba(35, 35, 35, 0.35);
                border-radius: 6px;
                background: #ffffff;
                box-shadow: 0 2px 10px rgba(0,0,0,0.22);
                color: #222;
                cursor: pointer;
                font-size: 13px;
                font-weight: 700;
                padding: 8px 10px;
            }
            .street-view-picker-button:hover,
            .street-view-picker-active {
                background: #fff8d8;
            }
            .street-view-picker-icon {
                position: relative;
                width: 16px;
                height: 22px;
                display: inline-block;
            }
            .street-view-picker-icon::before {
                content: "";
                position: absolute;
                left: 5px;
                top: 0;
                width: 7px;
                height: 7px;
                border-radius: 50%;
                background: #fbbc04;
                box-shadow: 0 0 0 1px #9a6f00 inset;
            }
            .street-view-picker-icon::after {
                content: "";
                position: absolute;
                left: 2px;
                top: 8px;
                width: 12px;
                height: 13px;
                border-radius: 7px 7px 4px 4px;
                background: #fbbc04;
                box-shadow: 0 0 0 1px #9a6f00 inset;
            }
        """

    def _add_single_route_customer_panel(
        self,
        route_map: folium.Map,
        route: Route,
        route_number: int,
        bus_color: str,
        marker_entries: List[Dict[str, object]],
    ) -> None:
        if not marker_entries:
            return

        rows = []
        panel_data = []
        route_stats_html = self._route_stats_html(route)
        for idx, entry in enumerate(marker_entries):
            number = int(entry["number"])
            name = html.escape(str(entry["name"]))
            customer_id = html.escape(str(entry["id"]))
            volume = float(entry["volume"])
            nav_url = html.escape(str(entry["navigation_url"]), quote=True)
            street_url = html.escape(str(entry.get(
                "street_view_url",
                self._street_view_url((float(entry["lat"]), float(entry["lng"]))),
            )), quote=True)
            rows.append(
                f'''
                <div class="route-client-card" role="button" tabindex="0" data-client-index="{idx}">
                    <div class="route-client-number" style="background:{html.escape(bus_color, quote=True)}">{number}</div>
                    <div class="route-client-main">
                        <div class="route-client-name">{name}</div>
                        <div class="route-client-meta">ID: {customer_id} · {volume:.2f} ст.</div>
                    </div>
                    <div class="route-client-actions">
                        <a class="route-client-nav" href="{nav_url}" target="_blank" rel="noopener">Навигация</a>
                        <a class="route-client-street" href="{street_url}" target="_blank" rel="noopener" title="Отвори Google Street View до клиента">
                            <span class="route-client-street-icon" aria-hidden="true"></span>
                            <span>Street View</span>
                        </a>
                    </div>
                </div>
                '''
            )
            panel_data.append(
                {
                    "markerName": entry["marker_name"],
                    "lat": entry["lat"],
                    "lng": entry["lng"],
                }
            )

        panel_json = json.dumps(panel_data, ensure_ascii=False)
        map_name = route_map.get_name()
        panel_html = f'''
        <style>
            .route-client-panel {{
                position: fixed;
                top: 10px;
                right: 10px;
                width: 340px;
                max-height: calc(100vh - 34px);
                overflow: auto;
                z-index: 9999;
                background: #ffffff;
                border: 1px solid rgba(35, 35, 35, 0.35);
                border-radius: 6px;
                box-shadow: 0 4px 18px rgba(0,0,0,0.22);
                padding: 10px;
                font-family: Arial, sans-serif;
                color: #222;
            }}
            .route-client-panel h4 {{
                margin: 0 0 8px;
                font-size: 15px;
                text-align: center;
            }}
            .route-client-summary {{
                margin: 0 0 8px;
                font-size: 12px;
                color: #555;
                text-align: center;
            }}
            .route-client-stats {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 6px;
                margin: 8px 0 10px;
                padding: 8px;
                background: #f8fafc;
                border: 1px solid #e1e6ee;
                border-radius: 6px;
            }}
            .route-client-stat {{
                display: flex;
                justify-content: space-between;
                gap: 8px;
                font-size: 12px;
                color: #1f2933;
            }}
            .route-client-stat b {{
                font-size: 12px;
            }}
            .route-client-card {{
                display: grid;
                grid-template-columns: 30px 1fr auto;
                gap: 8px;
                align-items: center;
                padding: 7px 6px;
                border-top: 1px solid #ececec;
                cursor: pointer;
            }}
            .route-client-card:hover, .route-client-card:focus {{
                background: #f4f7fb;
                outline: none;
            }}
            .route-client-number {{
                width: 26px;
                height: 26px;
                border-radius: 50%;
                color: #fff;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 700;
                border: 2px solid #fff;
                box-shadow: 0 1px 4px rgba(0,0,0,0.35);
            }}
            .route-client-name {{
                font-size: 12px;
                font-weight: 700;
                line-height: 1.25;
                overflow-wrap: anywhere;
            }}
            .route-client-meta {{
                font-size: 11px;
                color: #666;
                margin-top: 2px;
            }}
            .route-client-actions {{
                display: flex;
                flex-direction: column;
                gap: 5px;
                align-items: stretch;
            }}
            .route-client-nav, .route-client-street {{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                gap: 4px;
                padding: 5px 7px;
                border-radius: 4px;
                text-decoration: none;
                font-size: 11px;
                font-weight: 700;
                white-space: nowrap;
                text-align: center;
            }}
            .route-client-nav {{
                background: #1a73e8;
                color: #fff !important;
            }}
            .route-client-street {{
                background: #fbbc04;
                color: #1f1f1f !important;
                box-shadow: 0 0 0 1px #9a6f00 inset;
            }}
            .route-client-street-icon {{
                position: relative;
                display: inline-block;
                width: 9px;
                height: 13px;
                border-radius: 5px 5px 3px 3px;
                background: #1f1f1f;
                flex: 0 0 auto;
            }}
            .route-client-street-icon::before {{
                content: "";
                position: absolute;
                left: 2px;
                top: -5px;
                width: 5px;
                height: 5px;
                border-radius: 50%;
                background: #1f1f1f;
            }}
            @media (max-width: 700px) {{
                .route-client-panel {{
                    left: 8px;
                    right: 8px;
                    bottom: 8px;
                    top: auto;
                    width: auto;
                    max-height: 42vh;
                }}
            }}
        </style>
        <div class="route-client-panel">
            <h4>Клиенти - маршрут {route_number}</h4>
            {route_stats_html}
            {''.join(rows)}
        </div>
        <script>
            (function() {{
                const map = {map_name};
                const clients = {panel_json};

                function openClient(index) {{
                    const item = clients[index];
                    if (!item) return;
                    const marker = window[item.markerName];
                    if (!marker) return;
                    map.setView([item.lat, item.lng], Math.max(map.getZoom(), 15), {{ animate: true }});
                    marker.openPopup();
                }}

                document.querySelectorAll(".route-client-card").forEach((row) => {{
                    row.addEventListener("click", (event) => {{
                        if (event.target.closest("a")) return;
                        openClient(Number(row.dataset.clientIndex));
                    }});
                    row.addEventListener("keydown", (event) => {{
                        if (event.key === "Enter" || event.key === " ") {{
                            event.preventDefault();
                            openClient(Number(row.dataset.clientIndex));
                        }}
                    }});
                }});
            }})();
        </script>
        '''
        route_map.get_root().add_child(folium.Element(panel_html))

    def _add_single_route_customer_control(
        self,
        route_map: folium.Map,
        route: Route,
        route_number: int,
        bus_color: str,
        marker_entries: List[Dict[str, object]],
    ) -> None:
        if not marker_entries:
            return

        rows = []
        panel_data = []
        safe_color = html.escape(bus_color, quote=True)
        route_stats_html = self._route_stats_html(route)
        for idx, entry in enumerate(marker_entries):
            number = int(entry["number"])
            name = html.escape(str(entry["name"]))
            customer_id = html.escape(str(entry["id"]))
            volume = float(entry["volume"])
            nav_url = html.escape(str(entry["navigation_url"]), quote=True)
            street_url = html.escape(str(entry.get(
                "street_view_url",
                self._street_view_url((float(entry["lat"]), float(entry["lng"]))),
            )), quote=True)
            rows.append(
                f'''
                <div class="route-client-card" role="button" tabindex="0" data-client-index="{idx}">
                    <div class="route-client-number" style="background:{safe_color}">{number}</div>
                    <div class="route-client-main">
                        <div class="route-client-name">{name}</div>
                        <div class="route-client-meta">ID: {customer_id} &middot; {volume:.2f} ст.</div>
                    </div>
                    <div class="route-client-actions">
                        <a class="route-client-nav" href="{nav_url}" target="_blank" rel="noopener">Навигация</a>
                        <a class="route-client-street" href="{street_url}" target="_blank" rel="noopener" title="Отвори Google Street View до клиента">
                            <span class="route-client-street-icon" aria-hidden="true"></span>
                            <span>Street View</span>
                        </a>
                    </div>
                </div>
                '''
            )
            panel_data.append({
                "markerName": entry["marker_name"],
                "lat": entry["lat"],
                "lng": entry["lng"],
            })

        panel_html = f'''
            <div class="route-client-header">
                <h4>Клиенти - маршрут {route_number}</h4>
                <button type="button" class="route-client-close" title="Затвори">×</button>
            </div>
            {route_stats_html}
            {''.join(rows)}
        '''
        panel_css = '''
            .route-client-control {
                font-family: Arial, sans-serif;
            }
            .route-client-toggle {
                border: 1px solid rgba(35, 35, 35, 0.35);
                border-radius: 6px;
                background: #ffffff;
                box-shadow: 0 2px 10px rgba(0,0,0,0.22);
                color: #222;
                cursor: pointer;
                font-size: 13px;
                font-weight: 700;
                padding: 8px 11px;
            }
            .route-client-toggle:hover {
                background: #f4f7fb;
            }
            .route-client-toggle-hidden,
            .route-client-panel-hidden {
                display: none;
            }
            .route-client-panel {
                width: 340px;
                max-height: calc(100vh - 90px);
                overflow: auto;
                background: #ffffff;
                border: 1px solid rgba(35, 35, 35, 0.35);
                border-radius: 6px;
                box-shadow: 0 4px 18px rgba(0,0,0,0.22);
                padding: 10px;
                font-family: Arial, sans-serif;
                color: #222;
            }
            .route-client-header {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 8px;
                margin-bottom: 8px;
            }
            .route-client-header h4 {
                margin: 0;
                font-size: 15px;
            }
            .route-client-close {
                width: 26px;
                height: 26px;
                border: 0;
                border-radius: 4px;
                background: #f1f3f4;
                color: #333;
                cursor: pointer;
                font-size: 18px;
                line-height: 1;
                font-weight: 700;
            }
            .route-client-close:hover {
                background: #e4e7eb;
            }
            .route-client-summary {
                margin: 0 0 8px;
                font-size: 12px;
                color: #555;
                text-align: center;
            }
            .route-client-stats {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 6px;
                margin: 8px 0 10px;
                padding: 8px;
                background: #f8fafc;
                border: 1px solid #e1e6ee;
                border-radius: 6px;
            }
            .route-client-stat {
                display: flex;
                justify-content: space-between;
                gap: 8px;
                font-size: 12px;
                color: #1f2933;
            }
            .route-client-stat b {
                font-size: 12px;
            }
            .route-client-card {
                display: grid;
                grid-template-columns: 30px minmax(0, 1fr) auto;
                gap: 8px;
                align-items: center;
                padding: 7px 6px;
                border-top: 1px solid #ececec;
                cursor: pointer;
            }
            .route-client-card:hover, .route-client-card:focus {
                background: #f4f7fb;
                outline: none;
            }
            .route-client-number {
                width: 26px;
                height: 26px;
                border-radius: 50%;
                color: #fff;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 700;
                border: 2px solid #fff;
                box-shadow: 0 1px 4px rgba(0,0,0,0.35);
            }
            .route-client-name {
                font-size: 12px;
                font-weight: 700;
                line-height: 1.25;
                overflow-wrap: anywhere;
            }
            .route-client-meta {
                font-size: 11px;
                color: #666;
                margin-top: 2px;
            }
            .route-client-actions {
                display: flex;
                flex-direction: column;
                gap: 5px;
                align-items: stretch;
            }
            .route-client-nav, .route-client-street {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                gap: 4px;
                padding: 5px 7px;
                border-radius: 4px;
                text-decoration: none;
                font-size: 11px;
                font-weight: 700;
                white-space: nowrap;
                text-align: center;
            }
            .route-client-nav {
                background: #1a73e8;
                color: #fff !important;
            }
            .route-client-street {
                background: #fbbc04;
                color: #1f1f1f !important;
                box-shadow: 0 0 0 1px #9a6f00 inset;
            }
            .route-client-street-icon {
                position: relative;
                display: inline-block;
                width: 9px;
                height: 13px;
                border-radius: 5px 5px 3px 3px;
                background: #1f1f1f;
                flex: 0 0 auto;
            }
            .route-client-street-icon::before {
                content: "";
                position: absolute;
                left: 2px;
                top: -5px;
                width: 5px;
                height: 5px;
                border-radius: 50%;
                background: #1f1f1f;
            }
            @media (max-width: 700px) {
                .route-client-panel {
                    width: calc(100vw - 36px);
                    max-height: 42vh;
                }
            }
        '''
        route_map.add_child(SingleRouteCustomerPanel(panel_html, panel_data, panel_css))
    
    def create_single_route_map(self, route: Route, route_number: int,
                                depot_location: Tuple[float, float]) -> folium.Map:
        """Създава интерактивна карта за един отделен маршрут"""
        logger.info(f"Създавам карта за маршрут {route_number}")

        # Взимаме депото на маршрута
        route_depot = route.depot_location or depot_location

        # Център на картата - средна точка на клиентите или депото
        if route.customers:
            coords = [c.coordinates for c in route.customers if c.coordinates]
            if coords:
                avg_lat = sum(c[0] for c in coords) / len(coords)
                avg_lon = sum(c[1] for c in coords) / len(coords)
                center = (avg_lat, avg_lon)
            else:
                center = route_depot
        else:
            center = route_depot

        if self._use_google_maps():
            return self._build_google_html(
                f"Маршрут {route_number}",
                center,
                [route],
                [route_depot],
                single_route_number=route_number,
            )

        route_map = self._create_folium_map(center)
        route_map.add_child(StreetViewPicker(self._street_view_picker_css()))

        # Добавяме маркер за депото
        self._add_depot_markers(route_map, [route_depot])

        # Добавяме център зоната
        from config import get_config
        cfg = get_config()
        if cfg.locations.enable_center_zone_priority:
            self._add_center_zone_shape(route_map, cfg.locations)

        # Добавяме маршрута
        vehicle_settings = VEHICLE_SETTINGS.get(route.vehicle_type.value, {
            'color': 'gray', 'icon': 'circle', 'prefix': 'fa', 'name': 'Неизвестен'
        })
        bus_color = BUS_COLORS[(route_number - 1) % len(BUS_COLORS)]

        bus_layer = folium.FeatureGroup(
            name=f"\U0001f68c Автобус {route_number} ({len(route.customers)} клиента)")
        marker_entries = []

        # Маркери за клиентите
        for client_idx, customer in enumerate(route.customers):
            if customer.coordinates:
                client_number = client_idx + 1
                navigation_url = self._navigation_url(customer.coordinates)
                street_view_url = self._street_view_url(customer.coordinates)
                icon_html = f'''
                <div style="
                    background-color: {bus_color};
                    border: 3px solid white;
                    border-radius: 50%;
                    width: 30px;
                    height: 30px;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    font-weight: bold;
                    font-size: 14px;
                    color: white;
                    text-shadow: 1px 1px 1px rgba(0,0,0,0.7);
                ">{client_number}</div>
                '''
                popup_text = f"""
                <div style="font-family: Arial, sans-serif;">
                    <h4 style="margin: 0; color: {bus_color};">
                        Автобус {route_number} - {vehicle_settings['name']}
                    </h4>
                    <hr style="margin: 5px 0;">
                    <b>Клиент:</b> {customer.name}<br>
                    <b>ID:</b> {customer.id}<br>
                    <b>Ред в маршрута:</b> #{client_number}<br>
                    <b>Обем:</b> {customer.volume:.2f} ст.<br>
                    <b>Координати:</b> {customer.coordinates[0]:.6f}, {customer.coordinates[1]:.6f}<br>
                    {self._popup_action_buttons(navigation_url, street_view_url)}
                </div>
                """
                marker = folium.Marker(
                    customer.coordinates,
                    popup=folium.Popup(popup_text, max_width=300),
                    tooltip=f"#{client_number}: {customer.name}",
                    icon=folium.DivIcon(
                        html=icon_html,
                        icon_size=(30, 30),
                        icon_anchor=(15, 15),
                        popup_anchor=(0, -15)
                    )
                )
                marker.add_to(bus_layer)
                marker_entries.append({
                    "marker_name": marker.get_name(),
                    "number": client_number,
                    "name": customer.name,
                    "id": customer.id,
                    "volume": customer.volume,
                    "lat": customer.coordinates[0],
                    "lng": customer.coordinates[1],
                    "navigation_url": navigation_url,
                    "street_view_url": street_view_url,
                })

        # Линия на маршрута
        if route.customers:
            waypoints = [route_depot]
            for customer in route.customers:
                if customer.coordinates:
                    waypoints.append(customer.coordinates)
            waypoints.append(route_depot)

            if self.use_routing:
                engine_name = "Valhalla" if self.routing_engine and self.routing_engine.value == RoutingEngine.VALHALLA.value else "OSRM"
                try:
                    route_geometry = self._get_full_route_geometry(waypoints)
                    if len(route_geometry) > 2:
                        polyline = folium.PolyLine(
                            route_geometry, color=bus_color, weight=4, opacity=0.8
                        )
                    else:
                        polyline = folium.PolyLine(
                            waypoints, color=bus_color, weight=3, opacity=0.6, dash_array='5, 5'
                        )
                    polyline.add_to(bus_layer)
                    self._add_direction_arrows(polyline, bus_layer, bus_color)
                except Exception as e:
                    logger.warning(f"Грешка при маршрут геометрия за маршрут {route_number}: {e}")
                    polyline = folium.PolyLine(
                        waypoints, color=bus_color, weight=3, opacity=0.6, dash_array='5, 5'
                    )
                    polyline.add_to(bus_layer)
                    self._add_direction_arrows(polyline, bus_layer, bus_color)
            else:
                polyline = folium.PolyLine(
                    waypoints, color=bus_color, weight=3, opacity=0.8
                )
                polyline.add_to(bus_layer)
                self._add_direction_arrows(polyline, bus_layer, bus_color)

        bus_layer.add_to(route_map)
        self._add_single_route_customer_control(route_map, route, route_number, bus_color, marker_entries)

        # Легенда с информация за маршрута
        legend_html = f'''
        <div style="position: fixed;
                    top: 10px; left: 10px; width: 260px; height: auto;
                    background-color: white; border:2px solid grey; z-index:9999;
                    font-size:14px; padding: 10px; border-radius: 5px;
                    box-shadow: 0 0 15px rgba(0,0,0,0.2);">
        <h4 style="margin-top:0; margin-bottom:10px; text-align: center;">
            \U0001f68c Маршрут {route_number} - {vehicle_settings['name']}
        </h4>
        <p style="margin: 3px 0; font-size: 12px;">\U0001f4ca Клиенти: {len(route.customers)}</p>
        <p style="margin: 3px 0; font-size: 12px;">\U0001f4cf Разстояние: {route.total_distance_km:.1f} км</p>
        <p style="margin: 3px 0; font-size: 12px;">\u23f1 Време: {route.total_time_minutes:.0f} мин</p>
        <p style="margin: 3px 0; font-size: 12px;">\U0001f4e6 Обем: {route.total_volume:.1f} ст.</p>
        </div>
        '''
        route_map.get_root().add_child(folium.Element(legend_html))

        return route_map

    def save_map(self, route_map: folium.Map, file_path: Optional[str] = None) -> str:
        """Записва картата във файл"""
        file_path = _normalize_output_file_path(
            file_path or self.config.map_output_file,
            "interactive_map.html",
        )
        file_path = _append_run_date_to_filename(file_path, self.run_date)
        
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        route_map.save(file_path)
        
        logger.info(f"Интерактивна карта записана в {file_path}")
        return file_path


class ExcelExporter:
    """Експортър на Excel файлове"""
    
    def __init__(self, config: OutputConfig):
        self.config = config
        self.run_date = datetime.now().strftime("%Y-%m-%d")
    
    def export_all_to_single_excel(self, solution: CVRPSolution, warehouse_customers: List[Customer]) -> str:
        """Експортира всички данни в един Excel файл с отделни sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Създаваме основния файл
        file_path = os.path.join(self.config.excel_output_dir, "cvrp_report.xlsx")
        file_path = _append_run_date_to_filename(file_path, self.run_date)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        wb = Workbook()
        
        # Премахваме default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # 1. SHEET: Маршрути (Vehicle Routes)
        if solution.routes:
            self._create_routes_sheet(wb, solution)
        
        # 2. SHEET: Необслужени клиенти (Unserved Customers)
        if warehouse_customers:
            self._create_unserved_sheet(wb, warehouse_customers)
        
        # 3. SHEET: Обобщение (Summary)
        self._create_summary_sheet(wb, solution, warehouse_customers)
        
        # 4. SHEET: Статистики по автобуси (Vehicle Statistics)
        if solution.routes:
            self._create_vehicle_stats_sheet(wb, solution)
        
        # Записваме файла
        wb.save(file_path)
        logger.info(f"Общ Excel отчет записан в {file_path}")
        return file_path

    def _format_movement_direction(self, from_stop: str, to_stop: str) -> str:
        return f"{from_stop} -> {to_stop}"
    
    def _create_routes_sheet(self, wb, solution: CVRPSolution):
        """Създава sheet с маршрутите"""
        ws = wb.create_sheet("Маршрути")
        
        # Заглавни редове
        headers = [
            'Маршрут', 'Превозно средство', 'Ред в маршрута', 
            'Посока на движение', 'ID клиент', 'Име клиент', 'Номер поръчка', 'Обем (ст.)', 'GPS координати',
            'Разстояние до центъра (км)', 'Депо стартова точка',
            'Разстояние от предишен (км)', 'Накоплено разстояние (км)',
            'Време от предишен (мин)', 'Накоплено време (мин)',
            'Стартово време (мин)', 'Време с натрупване (мин)', 'Време с натрупване (чч:мм)'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни за маршрутите
        row = 2
        center_location = get_config().locations.center_location
        
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            
            # Изчисляваме стартово време за този тип превозно средство
            start_time_minutes = self._get_start_time_for_vehicle(route.vehicle_type)
            
            # Взимаме service time за този тип превозно средство
            vehicle_config = self._get_vehicle_config(route.vehicle_type)
            service_time_minutes = vehicle_config.service_time_minutes if vehicle_config else 15
            
            # Изчисляваме разстоянията и времената между клиентите
            cumulative_distance = 0
            cumulative_time = 0
            previous_customer_coords = route.depot_location  # Започваме от депото
            previous_stop_name = "Депо"
            
            for j, customer in enumerate(route.customers):
                # Изчисляваме разстоянието до центъра
                distance_to_center = self._calculate_distance_to_center(customer.coordinates, center_location) if customer.coordinates else 0.0
                
                # Изчисляваме разстоянието от предишния клиент
                distance_from_previous = self._calculate_distance_between_points(
                    previous_customer_coords, customer.coordinates
                ) if customer.coordinates else 0.0
                cumulative_distance += distance_from_previous
                
                # Изчисляваме времето от предишния клиент (приблизително)
                time_from_previous = self._calculate_time_between_points(
                    previous_customer_coords, customer.coordinates
                ) if customer.coordinates else 0.0
                
                # Добавяме service time за текущия клиент
                total_time_for_this_step = time_from_previous + service_time_minutes
                cumulative_time += total_time_for_this_step
                
                # Изчисляваме времето с натрупване (стартово време + натрупване)
                total_time_with_start = start_time_minutes + cumulative_time
                
                # Проверяваме дали клиентът е в център зоната
                is_in_center_zone = is_location_in_center_zone(customer.coordinates, get_config().locations)
                
                data = [
                    i + 1,  # Маршрут
                    vehicle_name,  # Превозно средство
                    j + 1,  # Ред в маршрута
                    self._format_movement_direction(previous_stop_name, customer.name),
                    customer.id,  # ID клиент
                    customer.name,  # Име клиент
                    customer.document,  # Номер поръчка
                    customer.volume,  # Обем
                    customer.original_gps_data,  # GPS
                    round(distance_to_center, 2),  # Разстояние до центъра
                    f"{route.depot_location[0]:.6f}, {route.depot_location[1]:.6f}",  # Депо
                    round(distance_from_previous, 2),  # Разстояние от предишен
                    round(cumulative_distance, 2),  # Накоплено разстояние
                    round(total_time_for_this_step, 1),  # Време от предишен + service time
                    round(cumulative_time, 1),  # Накоплено време
                    start_time_minutes,  # Стартово време (мин)
                    round(total_time_with_start, 1),  # Време с натрупване (мин)
                    self._format_time_hh_mm(int(total_time_with_start))  # Време с натрупване (чч:мм)
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
                previous_customer_coords = customer.coordinates
                previous_stop_name = customer.name
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_unserved_sheet(self, wb, warehouse_customers: List[Customer]):
        """Създава sheet с необслужените клиенти"""
        ws = wb.create_sheet("Необслужени клиенти")
        
        headers = [
            'ID', 'Име', 'Обем (ст.)', 'GPS координати', 
            'Latitude', 'Longitude', 'Разстояние до центъра (км)'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни
        center_location = get_config().locations.center_location
        row = 2
        
        for customer in warehouse_customers:
            distance_to_center = self._calculate_distance_to_center(customer.coordinates, center_location)
            
            data = [
                customer.id,
                customer.name,
                customer.volume,
                customer.original_gps_data,
                customer.coordinates[0] if customer.coordinates else '',
                customer.coordinates[1] if customer.coordinates else '',
                round(distance_to_center, 2)
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb, solution: CVRPSolution, warehouse_customers: List[Customer]):
        """Създава sheet с обобщение"""
        ws = wb.create_sheet("Обобщение")
        
        # Стилове
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Заглавие
        ws['A1'] = "CVRP ОТЧЕТ - ОБОБЩЕНИЕ"
        ws['A1'].font = title_font
        
        # Основни статистики
        row = 3
        stats = [
            ("Общо клиенти", len(solution.routes) + len(warehouse_customers)),
            ("Обслужени клиенти", sum(len(route.customers) for route in solution.routes)),
            ("Необслужени клиенти", len(warehouse_customers)),
            ("Брой маршрути", len(solution.routes)),
            ("Общо разстояние (км)", round(solution.total_distance_km, 2)),
            ("Общо време (мин)", round(solution.total_time_minutes, 2)),
            ("Общ обем (ст.)", round(sum(route.total_volume for route in solution.routes), 2))
        ]
        
        for stat_name, stat_value in stats:
            ws[f'A{row}'] = stat_name
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = stat_value
            row += 1
        
        # Информация за стартови времена
        row += 2
        ws[f'A{row}'] = "СТАРТОВИ ВРЕМЕНА ПО ТИП АВТОБУС"
        ws[f'A{row}'].font = title_font
        row += 1
        
        # Заглавни редове за стартови времена
        start_time_headers = ['Тип автобус', 'Стартово време (мин)', 'Стартово време (чч:мм)']
        for col, header in enumerate(start_time_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # Данни за стартови времена
        vehicle_types_seen = set()
        for route in solution.routes:
            if route.vehicle_type.value not in vehicle_types_seen:
                vehicle_types_seen.add(route.vehicle_type.value)
                vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', route.vehicle_type.value)
                start_time_minutes = self._get_start_time_for_vehicle(route.vehicle_type)
                
                data = [
                    vehicle_name,
                    start_time_minutes,
                    self._format_time_hh_mm(start_time_minutes)
                ]
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
        
        # Статистики по тип автобус
        row += 2
        ws[f'A{row}'] = "СТАТИСТИКИ ПО ТИП АВТОБУС"
        ws[f'A{row}'].font = title_font
        row += 1
        
        vehicle_stats = {}
        for route in solution.routes:
            vehicle_type = route.vehicle_type.value
            if vehicle_type not in vehicle_stats:
                vehicle_stats[vehicle_type] = {
                    'count': 0, 'distance': 0, 'volume': 0, 'customers': 0
                }
            vehicle_stats[vehicle_type]['count'] += 1
            vehicle_stats[vehicle_type]['distance'] += route.total_distance_km
            vehicle_stats[vehicle_type]['volume'] += route.total_volume
            vehicle_stats[vehicle_type]['customers'] += len(route.customers)
        
        # Заглавни редове за статистики
        headers = ['Тип автобус', 'Брой маршрути', 'Общо разстояние (км)', 'Общ обем (ст.)', 'Общо клиенти']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # Данни за статистики
        for vehicle_type, stats in vehicle_stats.items():
            vehicle_name = VEHICLE_SETTINGS.get(vehicle_type, {}).get('name', vehicle_type)
            data = [
                vehicle_name,
                stats['count'],
                round(stats['distance'], 2),
                round(stats['volume'], 2),
                stats['customers']
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_vehicle_stats_sheet(self, wb, solution: CVRPSolution):
        """Създава sheet със статистики по отделни автобуси"""
        ws = wb.create_sheet("Статистики по автобуси")
        
        headers = [
            'Маршрут', 'Тип автобус', 'Брой клиенти', 'Общ обем (ст.)',
            'Разстояние (км)', 'Време (мин)', 'Капацитет използване (%)',
            'Средно разстояние до центъра (км)', 'Депо стартова точка', 'Стартово време (чч:мм)'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни
        center_location = get_config().locations.center_location
        row = 2
        
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            
            # Изчисляваме средното разстояние до центъра
            distances_to_center = []
            for customer in route.customers:
                distance = self._calculate_distance_to_center(customer.coordinates, center_location)
                distances_to_center.append(distance)
            avg_distance_to_center = sum(distances_to_center) / len(distances_to_center) if distances_to_center else 0
            
            # Капацитет използване (трябва да вземем capacity от config)
            vehicle_config = self._get_vehicle_config(route.vehicle_type)
            capacity_usage = 0
            if vehicle_config and vehicle_config.capacity > 0:
                capacity_usage = (route.total_volume / vehicle_config.capacity * 100)
            
            # Изчисляваме стартово време за този тип превозно средство
            start_time_minutes = self._get_start_time_for_vehicle(route.vehicle_type)
            
            data = [
                i + 1,  # Маршрут
                vehicle_name,  # Тип автобус
                len(route.customers),  # Брой клиенти
                round(route.total_volume, 2),  # Общ обем
                round(route.total_distance_km, 2),  # Разстояние
                round(route.total_time_minutes, 2),  # Време
                round(capacity_usage, 1),  # Капацитет използване
                round(avg_distance_to_center, 2),  # Средно разстояние до центъра
                f"{route.depot_location[0]:.6f}, {route.depot_location[1]:.6f}",  # Депо
                self._format_time_hh_mm(start_time_minutes) # Стартово време (чч:мм)
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _calculate_distance_to_center(self, coordinates: Optional[Tuple[float, float]], center_location: Tuple[float, float]) -> float:
        """Изчислява разстоянието до центъра в км"""
        if not coordinates or not center_location:
            return 0.0
        
        from math import radians, sin, cos, sqrt, atan2
        R = 6371  # Earth radius in km
        
        lat1, lon1 = radians(coordinates[0]), radians(coordinates[1])
        lat2, lon2 = radians(center_location[0]), radians(center_location[1])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def _calculate_distance_between_points(self, point1: Optional[Tuple[float, float]], point2: Optional[Tuple[float, float]]) -> float:
        """Изчислява разстоянието между две точки в км"""
        if not point1 or not point2:
            return 0.0
        
        from math import radians, sin, cos, sqrt, atan2
        R = 6371  # Earth radius in km
        
        lat1, lon1 = radians(point1[0]), radians(point1[1])
        lat2, lon2 = radians(point2[0]), radians(point2[1])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def _calculate_time_between_points(self, point1: Optional[Tuple[float, float]], point2: Optional[Tuple[float, float]]) -> float:
        """Изчислява времето за пътуване между две точки в минути (приблизително)"""
        if not point1 or not point2:
            return 0.0
        
        distance_km = self._calculate_distance_between_points(point1, point2)
        # Приблизително време за пътуване: 2 минути на км (градски транспорт)
        # Това включва спирачки, светофари, задръствания и т.н.
        return distance_km * 2

    def _get_vehicle_config(self, vehicle_type):
        """Връща конфигурацията за даден тип превозно средство"""
        from config import get_config
        vehicle_configs = get_config().vehicles
        
        if vehicle_configs:
            for config in vehicle_configs:
                if config.vehicle_type == vehicle_type:
                    return config
        return None
    
    def _get_start_time_for_vehicle(self, vehicle_type) -> int:
        """Връща стартово време в минути за даден тип превозно средство"""
        vehicle_config = self._get_vehicle_config(vehicle_type)
        if vehicle_config and hasattr(vehicle_config, 'start_time_minutes'):
            return vehicle_config.start_time_minutes
        else:
            # Използваме глобалното стартово време от конфигурацията
            from config import get_config
            return get_config().cvrp.global_start_time_minutes
    
    def _format_time_hh_mm(self, total_minutes: int) -> str:
        """Форматира време в минути като чч:мм"""
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    
    def export_warehouse_orders(self, warehouse_customers: List[Customer]) -> str:
        """Експортира заявките в склада (за съвместимост)"""
        if not warehouse_customers:
            logger.info("Няма заявки за експорт в склада")
            return ""
        
        file_path = os.path.join(self.config.excel_output_dir, self.config.warehouse_excel_file)
        file_path = _append_run_date_to_filename(file_path, self.run_date)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        data = []
        for customer in warehouse_customers:
            data.append({
                'ID': customer.id,
                'Име': customer.name,
                'Обем (ст.)': customer.volume,
                'GPS координати': customer.original_gps_data,
                'Latitude': customer.coordinates[0] if customer.coordinates else '',
                'Longitude': customer.coordinates[1] if customer.coordinates else ''
            })
        
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)
        
        logger.info(f"Складови заявки експортирани в {file_path}")
        return file_path
    
    def export_vehicle_routes(self, solution: CVRPSolution) -> str:
        """Експортира маршрутите на превозните средства (за съвместимост)"""
        file_path = os.path.join(self.config.excel_output_dir, self.config.routes_excel_file)
        file_path = _append_run_date_to_filename(file_path, self.run_date)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        data = []
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            previous_stop_name = "Депо"
            for j, customer in enumerate(route.customers):
                data.append({
                    'Маршрут': i + 1,
                    'Превозно средство': vehicle_name,
                    'Ред в маршрута': j + 1,
                    'Посока на движение': self._format_movement_direction(previous_stop_name, customer.name),
                    'ID клиент': customer.id,
                    'Име клиент': customer.name,
                    'Номер поръчка': customer.document,
                    'Обем (ст.)': customer.volume,
                    'GPS': customer.original_gps_data
                })
                previous_stop_name = customer.name
        
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        logger.info(f"Маршрути експортирани в {file_path}")
        return file_path

    def export_routes_csv(self, solution: CVRPSolution) -> str:
        """Експортира маршрутите като CSV файл (разделен по запетаи)"""
        import csv
        
        file_path = _normalize_output_file_path(self.config.csv_output_file, "routes.csv")
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        headers = [
            'Маршрут', 'Превозно средство', 'Ред в маршрута',
            'Посока на движение', 'ID клиент', 'Име клиент', 'Номер поръчка', 'Обем (ст.)', 'GPS координати',
            'Разстояние до центъра (км)', 'Депо стартова точка',
            'Разстояние от предишен (км)', 'Накоплено разстояние (км)',
            'Време от предишен (мин)', 'Накоплено време (мин)',
            'Стартово време (мин)', 'Време с натрупване (мин)', 'Време с натрупване (чч:мм)'
        ]
        
        center_location = get_config().locations.center_location
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=',')
            writer.writerow(headers)
            
            for i, route in enumerate(solution.routes):
                vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
                start_time_minutes = self._get_start_time_for_vehicle(route.vehicle_type)
                vehicle_config = self._get_vehicle_config(route.vehicle_type)
                service_time_minutes = vehicle_config.service_time_minutes if vehicle_config else 15
                
                cumulative_distance = 0
                cumulative_time = 0
                previous_customer_coords = route.depot_location
                previous_stop_name = "Депо"
                
                for j, customer in enumerate(route.customers):
                    distance_to_center = self._calculate_distance_to_center(
                        customer.coordinates, center_location) if customer.coordinates else 0.0
                    distance_from_previous = self._calculate_distance_between_points(
                        previous_customer_coords, customer.coordinates) if customer.coordinates else 0.0
                    cumulative_distance += distance_from_previous
                    time_from_previous = self._calculate_time_between_points(
                        previous_customer_coords, customer.coordinates) if customer.coordinates else 0.0
                    total_time_for_this_step = time_from_previous + service_time_minutes
                    cumulative_time += total_time_for_this_step
                    total_time_with_start = start_time_minutes + cumulative_time
                    
                    row = [
                        i + 1,
                        vehicle_name,
                        j + 1,
                        self._format_movement_direction(previous_stop_name, customer.name),
                        customer.id,
                        customer.name,
                        customer.document,
                        customer.volume,
                        customer.original_gps_data,
                        round(distance_to_center, 2),
                        f"{route.depot_location[0]:.6f}, {route.depot_location[1]:.6f}",
                        round(distance_from_previous, 2),
                        round(cumulative_distance, 2),
                        round(total_time_for_this_step, 1),
                        round(cumulative_time, 1),
                        start_time_minutes,
                        round(total_time_with_start, 1),
                        self._format_time_hh_mm(int(total_time_with_start))
                    ]
                    writer.writerow(row)
                    previous_customer_coords = customer.coordinates
                    previous_stop_name = customer.name
        
        logger.info(f"CSV маршрути експортирани в {file_path}")
        return file_path


class ChartGenerator:
    """Генератор на графики за анализ на маршрутите"""

    def __init__(self, config: OutputConfig):
        self.config = config

    def generate_all_charts(self, solution: CVRPSolution,
                            warehouse_customers: List[Customer]) -> Dict[str, str]:
        """Генерира всички графики и връща пътищата до файловете"""
        if not MATPLOTLIB_AVAILABLE:
            logger.warning("matplotlib не е инсталиран — графиките се пропускат")
            return {}

        os.makedirs(self.config.charts_output_dir, exist_ok=True)
        output_files = {}

        try:
            path = self._generate_efficiency_chart(solution, warehouse_customers)
            if path:
                output_files['efficiency_chart'] = path
        except Exception as e:
            logger.error(f"Грешка при генериране на efficiency chart: {e}")

        try:
            path = self._generate_route_comparison_chart(solution)
            if path:
                output_files['route_comparison_chart'] = path
        except Exception as e:
            logger.error(f"Грешка при генериране на route comparison chart: {e}")

        try:
            path = self._generate_volume_distribution_chart(solution)
            if path:
                output_files['volume_distribution_chart'] = path
        except Exception as e:
            logger.error(f"Грешка при генериране на volume distribution chart: {e}")

        logger.info(f"Генерирани {len(output_files)} графики в {self.config.charts_output_dir}")
        return output_files

    def _generate_efficiency_chart(self, solution: CVRPSolution,
                                    warehouse_customers: List[Customer]) -> Optional[str]:
        """Графика с анализ на ефективността — капацитет, време, обслужени/необслужени"""
        if not solution.routes:
            return None

        fig, axes = plt.subplots(1, 3, figsize=(18, 6))
        fig.suptitle('Анализ на ефективността', fontsize=16, fontweight='bold')

        # 1. Капацитет използване по маршрут
        ax1 = axes[0]
        labels = []
        usages = []
        for i, route in enumerate(solution.routes):
            vc = self._get_vehicle_config(route.vehicle_type)
            cap = vc.capacity if vc else 1
            usage = (route.total_volume / cap * 100) if cap > 0 else 0
            labels.append(f"М{i+1}")
            usages.append(usage)

        colors = [BUS_COLORS[i % len(BUS_COLORS)] for i in range(len(labels))]
        bars = ax1.bar(labels, usages, color=colors, edgecolor='white')
        ax1.axhline(y=100, color='red', linestyle='--', linewidth=1, label='100%')
        ax1.set_ylabel('Използване (%)')
        ax1.set_title('Капацитет по маршрут')
        ax1.set_ylim(0, max(usages + [100]) * 1.15)
        for bar, val in zip(bars, usages):
            ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                     f'{val:.0f}%', ha='center', va='bottom', fontsize=8)

        # 2. Обслужени vs Необслужени клиенти (pie)
        ax2 = axes[1]
        served = sum(len(r.customers) for r in solution.routes)
        unserved = len(warehouse_customers) + len(solution.dropped_customers)
        if served + unserved > 0:
            ax2.pie([served, unserved],
                    labels=[f'Обслужени ({served})', f'Необслужени ({unserved})'],
                    autopct='%1.1f%%', colors=['#2ecc71', '#e74c3c'],
                    startangle=90, textprops={'fontsize': 10})
        ax2.set_title('Обслужени / Необслужени')

        # 3. Обобщена лента — разстояние, време, обем
        ax3 = axes[2]
        total_dist = solution.total_distance_km
        total_time = solution.total_time_minutes
        total_vol = solution.total_served_volume
        cats = ['Разстояние\n(км)', 'Време\n(мин)', 'Обем\n(ст.)']
        vals = [total_dist, total_time, total_vol]
        bar_colors = ['#3498db', '#e67e22', '#9b59b6']
        bars3 = ax3.bar(cats, vals, color=bar_colors, edgecolor='white')
        for bar, val in zip(bars3, vals):
            ax3.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                     f'{val:.1f}', ha='center', va='bottom', fontsize=9)
        ax3.set_title('Обобщени показатели')
        ax3.set_ylabel('Стойност')

        plt.tight_layout(rect=[0, 0, 1, 0.93])
        file_path = os.path.join(self.config.charts_output_dir, self.config.efficiency_chart_file)
        fig.savefig(file_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        logger.info(f"Графика записана: {file_path}")
        return file_path

    def _generate_route_comparison_chart(self, solution: CVRPSolution) -> Optional[str]:
        """Сравнение на маршрутите — разстояние, време, обем за всеки"""
        if not solution.routes:
            return None

        n = len(solution.routes)
        labels = [f"М{i+1}" for i in range(n)]
        distances = [r.total_distance_km for r in solution.routes]
        times = [r.total_time_minutes for r in solution.routes]
        volumes = [r.total_volume for r in solution.routes]
        clients = [len(r.customers) for r in solution.routes]

        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('Сравнение на маршрутите', fontsize=16, fontweight='bold')

        colors = [BUS_COLORS[i % len(BUS_COLORS)] for i in range(n)]

        # Разстояние
        ax = axes[0, 0]
        ax.bar(labels, distances, color=colors, edgecolor='white')
        ax.set_title('Разстояние (км)')
        ax.set_ylabel('км')
        for i, v in enumerate(distances):
            ax.text(i, v + 0.3, f'{v:.1f}', ha='center', fontsize=8)

        # Време
        ax = axes[0, 1]
        ax.bar(labels, times, color=colors, edgecolor='white')
        ax.set_title('Време (мин)')
        ax.set_ylabel('мин')
        for i, v in enumerate(times):
            ax.text(i, v + 0.3, f'{v:.0f}', ha='center', fontsize=8)

        # Обем
        ax = axes[1, 0]
        ax.bar(labels, volumes, color=colors, edgecolor='white')
        ax.set_title('Обем (ст.)')
        ax.set_ylabel('ст.')
        for i, v in enumerate(volumes):
            ax.text(i, v + 0.3, f'{v:.1f}', ha='center', fontsize=8)

        # Брой клиенти
        ax = axes[1, 1]
        ax.bar(labels, clients, color=colors, edgecolor='white')
        ax.set_title('Брой клиенти')
        ax.set_ylabel('бр.')
        for i, v in enumerate(clients):
            ax.text(i, v + 0.1, str(v), ha='center', fontsize=8)

        plt.tight_layout(rect=[0, 0, 1, 0.93])
        file_path = os.path.join(self.config.charts_output_dir, self.config.route_comparison_file)
        fig.savefig(file_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        logger.info(f"Графика записана: {file_path}")
        return file_path

    def _generate_volume_distribution_chart(self, solution: CVRPSolution) -> Optional[str]:
        """Разпределение на обемите по клиенти (хистограма + по маршрут)"""
        if not solution.routes:
            return None

        all_volumes = []
        route_labels = []
        route_volumes = []
        for i, route in enumerate(solution.routes):
            vols = [c.volume for c in route.customers]
            all_volumes.extend(vols)
            route_labels.append(f"М{i+1}")
            route_volumes.append(vols)

        if not all_volumes:
            return None

        fig, axes = plt.subplots(1, 2, figsize=(14, 6))
        fig.suptitle('Разпределение на обемите', fontsize=16, fontweight='bold')

        # 1. Хистограма на всички обеми
        ax1 = axes[0]
        ax1.hist(all_volumes, bins=min(20, len(set(all_volumes))), color='#3498db',
                 edgecolor='white', alpha=0.85)
        ax1.set_xlabel('Обем (ст.)')
        ax1.set_ylabel('Брой клиенти')
        ax1.set_title('Разпределение на обемите')
        mean_vol = sum(all_volumes) / len(all_volumes)
        ax1.axvline(mean_vol, color='red', linestyle='--', label=f'Средно: {mean_vol:.1f}')
        ax1.legend()

        # 2. Box plot по маршрут
        ax2 = axes[1]
        if route_volumes and any(route_volumes):
            bp = ax2.boxplot(route_volumes, labels=route_labels, patch_artist=True)
            for i, box in enumerate(bp['boxes']):
                box.set_facecolor(BUS_COLORS[i % len(BUS_COLORS)])
                box.set_alpha(0.7)
        ax2.set_xlabel('Маршрут')
        ax2.set_ylabel('Обем (ст.)')
        ax2.set_title('Обеми по маршрут')

        plt.tight_layout(rect=[0, 0, 1, 0.93])
        file_path = os.path.join(self.config.charts_output_dir, self.config.volume_distribution_file)
        fig.savefig(file_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        logger.info(f"Графика записана: {file_path}")
        return file_path

    def _get_vehicle_config(self, vehicle_type):
        """Връща конфигурацията за даден тип превозно средство"""
        vehicle_configs = get_config().vehicles
        if vehicle_configs:
            for config in vehicle_configs:
                if config.vehicle_type == vehicle_type:
                    return config
        return None


class OutputHandler:
    """Главен клас за управление на изходните данни"""
    
    def __init__(self, config: Optional[OutputConfig] = None):
        self.config = config or get_config().output
        self.excel_exporter = ExcelExporter(self.config)
    
    def generate_all_outputs(self, solution: CVRPSolution, 
                           warehouse_allocation: WarehouseAllocation,
                           depot_location: Tuple[float, float]) -> Dict[str, str]:
        """Генерира всички изходни файлове и връща речник с пътищата до тях"""
        logger.info("Започвам генериране на изходни файлове")
        output_files = {}

        # 1. Интерактивна карта (обща)
        if self.config.enable_interactive_map:
            map_gen = InteractiveMapGenerator(self.config)
            route_map = map_gen.create_map(solution, warehouse_allocation, depot_location)
            map_file = map_gen.save_map(route_map)
            output_files['map'] = map_file

            # 1.1. Отделни HTML карти за всеки маршрут
            routes_dir = self.config.routes_output_dir
            os.makedirs(routes_dir, exist_ok=True)
            for idx, route in enumerate(solution.routes):
                route_number = idx + 1
                single_map = map_gen.create_single_route_map(route, route_number, depot_location)
                route_file = os.path.join(routes_dir, f"route_{route_number}.html")
                saved_route_file = map_gen.save_map(single_map, route_file)
                output_files[f'route_map_{route_number}'] = saved_route_file
            logger.info(f"Генерирани {len(solution.routes)} отделни HTML карти в {routes_dir}")
        
        # 2. Обединяване на всички необслужени клиенти
        all_unserviced_customers = warehouse_allocation.warehouse_customers + solution.dropped_customers
        
        # 3. Експорт в един общ Excel файл с отделни sheets
        if solution.routes or all_unserviced_customers:
            excel_file = self.excel_exporter.export_all_to_single_excel(solution, all_unserviced_customers)
            if excel_file:
                output_files['excel_report'] = excel_file
        
        # 4. CSV файл с маршрутите
        if solution.routes:
            csv_file = self.excel_exporter.export_routes_csv(solution)
            if csv_file:
                output_files['csv_routes'] = csv_file
        
        # 5. Графики (charts)
        try:
            chart_gen = ChartGenerator(self.config)
            chart_files = chart_gen.generate_all_charts(solution, all_unserviced_customers)
            output_files.update(chart_files)
        except Exception as e:
            logger.error(f"Грешка при генериране на графики: {e}")
        
        logger.info(f"Генерирани {len(output_files)} изходни файла")
        return output_files


# Удобна функция
def generate_outputs(solution: CVRPSolution, warehouse_allocation: WarehouseAllocation,
                   depot_location: Tuple[float, float]) -> Dict[str, str]:
    """Удобна функция за генериране на всички изходи"""
    handler = OutputHandler()
    return handler.generate_all_outputs(solution, warehouse_allocation, depot_location) 
